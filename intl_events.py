"""
VFL 2026 International Events — Pricing & Team Building.

Step 1: Generate prices for Kickoff 2026 and Santiago 2026
Step 2: Evaluate pricing accuracy (generated vs manual)
Step 3: Build optimal Kickoff 2026 team
Step 4: Build Santiago 2026 teams (3 GWs with transfers)
Step 5: Create Excel output files

Usage:
    python intl_events.py
"""
import pandas as pd
import numpy as np
import os
import sys
import copy
import warnings
from datetime import datetime

warnings.filterwarnings("ignore", category=FutureWarning)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DIR = os.path.dirname(os.path.abspath(__file__))

from pricing_algorithms import (
    CHINA_TEAMS, STAGE_ORDER,
    compute_team_win_rates, load_pickrate_data,
    compute_player_pickrates, compute_team_popularity,
)

# Extended game order including all Kickoff/Santiago game codes
GAME_ORDER = {
    "G1": 0, "G2": 1, "G3": 2, "G4": 3, "G5": 4,
    "SR1": 5, "SR2": 6, "SR3": 7,
    "UR1": 10, "UR2": 11, "UR3": 12,
    "UQF": 13, "USF": 14, "UF": 15,
    "MR1": 16, "MR2": 17, "MR3": 18, "MR4": 19, "MF": 20,
    "LR1": 21, "LR2": 22, "LR3": 23, "LR4": 24, "LR5": 25,
    "LF": 26, "GF": 27,
}
from vfl_2026_v3 import (
    _read_csv_safe, _normalize_team_name,
    compute_player_event_ppms, dampen_outlier_events,
    compute_weighted_ppm, get_pickrate_signal, get_uncertainty_flag,
    TEAM_HISTORY,
)

RNG = np.random.default_rng(2026)

# Player name normalization for encoding mismatches
PLAYER_NAME_FIXES = {
    "Ros\x82": "Rose",
    "Rosé": "Rose",
    "LEVIATÁN": "LEVIATAN",
    "LEVIAT\xb5N": "LEVIATAN",
}


def _normalize_player_name(name):
    """Normalize player names for consistent matching."""
    if not isinstance(name, str):
        return name
    return PLAYER_NAME_FIXES.get(name, name)

# ============================================================================
#  CONSTANTS
# ============================================================================

# International event VP range (from actual price lists: 5.0-15.0)
INTL_VP_MIN, INTL_VP_MAX = 5.0, 15.0

# Kickoff
KICKOFF_BUDGET = 100
KICKOFF_SQUAD = 11
KICKOFF_BEST_N = 3  # best 3 games count

# Santiago
SANTIAGO_BUDGET = 50
SANTIAGO_SQUAD = 6
SANTIAGO_BEST_N = 2
SANTIAGO_MAX_TRANSFERS = 2

# Santiago Swiss teams (8)
SANTIAGO_SWISS_TEAMS = [
    "Gentle Mates", "Team Liquid", "T1", "Paper Rex",
    "G2 Esports", "NRG Esports", "Xi Lai Gaming", "EDward Gaming",
]

# China teams in international events (only the 3 that actually compete)
CHINA_TEAMS_INTL = {
    "Xi Lai Gaming", "EDward Gaming", "Edward Gaming", "All Gamers",
}

# Role slot configs
KICKOFF_ROLE_SLOTS = {"D": 2, "C": 2, "I": 2, "S": 2}
KICKOFF_WC_SLOTS = 3
SANTIAGO_ROLE_SLOTS = {"D": 1, "C": 1, "I": 1, "S": 1}
SANTIAGO_WC_SLOTS = 2

# S-curve for international (5.5-13.5 range)
INTL_TARGET_QUANTILES = np.array([0.00, 0.05, 0.10, 0.20, 0.30, 0.40, 0.50,
                                   0.60, 0.70, 0.80, 0.90, 0.95, 1.00])
INTL_TARGET_VPS = np.array([5.5, 6.0, 6.5, 7.0, 7.5, 8.0, 8.5,
                             9.0, 9.5, 10.5, 11.5, 12.5, 13.5])


# ============================================================================
#  DATA LOADING
# ============================================================================

def load_csv(year):
    """Load a single year CSV."""
    path = os.path.join(DIR, f"{year} VFL.csv")
    df = _read_csv_safe(path)
    if "Team" not in df.columns:
        cols = list(df.columns)
        if cols[0].startswith("Unnamed"):
            cols[0], cols[1] = "Team", "Player"
            df.columns = cols
    df["Year"] = year
    df["Team"] = df["Team"].apply(_normalize_team_name)
    df["Player"] = df["Player"].apply(_normalize_player_name)
    return df


def load_data_up_to(max_year, max_stage=None, include_china=False):
    """Load data with temporal boundary. If max_stage set, only include up to
    that stage in max_year."""
    frames = []
    for year in [2024, 2025, 2026]:
        if year > max_year:
            break
        df = load_csv(year)
        if year == max_year and max_stage is not None:
            stage_order = {"Kickoff": 0, "Santiago": 1, "Stage 1": 2,
                          "Stage 2": 3, "Champions": 4}
            max_ord = stage_order.get(max_stage, 99)
            df = df[df["Stage"].map(lambda s: stage_order.get(s, 99) <= max_ord)]
        frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    if not include_china:
        combined = combined[~combined["Team"].isin(CHINA_TEAMS)].copy()
    return combined


def get_event_players(csv_2026, stage, include_china=True):
    """Get player pool for a specific event from 2026 CSV."""
    mask = csv_2026["Stage"] == stage
    players = csv_2026.loc[mask, ["Team", "Player"]].drop_duplicates(subset="Player")
    if not include_china:
        players = players[~players["Team"].isin(CHINA_TEAMS)]
    return players


def load_manual_intl_prices(event):
    """Load manual prices from dedicated price CSV files.
    event: 'kickoff' or 'santiago'
    """
    filename = f"{event}_prices_2026.csv"
    path = os.path.join(DIR, filename)
    df = _read_csv_safe(path)
    df["Team"] = df["Team"].apply(_normalize_team_name)
    df = df.rename(columns={"Price": "manual_vp"})
    return df


def get_player_actual_game_pts(csv_2026, stage, week=None):
    """Get per-game points for each player in an event.
    Returns dict: player -> list of (game, pts, ppm)"""
    mask = (csv_2026["Stage"] == stage) & (csv_2026["P?"] == 1)
    if week:
        mask = mask & (csv_2026["Wk"] == week)
    data = csv_2026[mask].copy()

    result = {}
    for _, row in data.iterrows():
        player = row["Player"]
        if player not in result:
            result[player] = []
        result[player].append({
            "game": row["Game"],
            "pts": row["Pts"],
            "ppm": row["PPM"],
            "week": row["Wk"],
        })
    return result


def compute_best_n_total(game_list, n_best=2):
    """Sum of best N games from a list of game dicts."""
    if not game_list:
        return 0.0
    pts_list = sorted([g["pts"] for g in game_list], reverse=True)
    return sum(pts_list[:n_best])


# ============================================================================
#  STEP 1: GENERATE INTERNATIONAL PRICES
# ============================================================================

def get_intl_year_weights(games_recent):
    """Year weights for international pricing. 'recent' = the most recent year of data."""
    if games_recent >= 8:
        return {2024: 0.15, 2025: 0.35, 2026: 1.0}
    elif games_recent >= 4:
        return {2024: 0.35, 2025: 0.55, 2026: 0.85}
    elif games_recent >= 1:
        return {2024: 0.50, 2025: 0.70, 2026: 0.65}
    else:
        return {2024: 0.60, 2025: 0.80, 2026: 0.0}


def generate_intl_prices(all_data, player_pool, pickrate_dict, team_pop,
                          team_wr, target_mean, vp_min=5.5, vp_max=13.5,
                          role_map=None):
    """Generate prices for an international event.

    Args:
        all_data: historical data (temporal boundary already applied)
        player_pool: DataFrame with Player, Team columns
        pickrate_dict: player pickrate data
        team_pop: team brand popularity
        team_wr: team win rates
        target_mean: target average VP
        vp_min/vp_max: VP bounds
        role_map: dict player_lower -> role letter
    """
    # Compute event PPMs from available data
    event_ppms = compute_player_event_ppms(all_data)

    played = all_data[all_data["P?"] == 1]
    global_mean_ppm = played["PPM"].mean() if len(played) > 0 else 3.5

    # Team avg PPM for newcomers
    team_avg_ppm = {}
    for team_name, grp in played.groupby("Team"):
        team_avg_ppm[team_name] = grp["PPM"].mean()

    avg_tw = np.mean(list(team_wr.values())) if team_wr else 0.5

    # S-curve targets scaled to vp range
    target_quantiles = INTL_TARGET_QUANTILES
    target_vps = np.linspace(vp_min, vp_max, len(target_quantiles))
    # Shape it with S-curve (compress extremes, expand middle)
    target_vps = np.array([vp_min, vp_min+0.5, vp_min+1.0, vp_min+1.5,
                           vp_min+2.0, vp_min+2.5, target_mean,
                           target_mean+0.5, target_mean+1.0,
                           target_mean+2.0, target_mean+3.0,
                           vp_max-1.0, vp_max])

    records = []
    for _, row in player_pool.iterrows():
        player = row["Player"]
        team = row["Team"]
        pl = player.lower()
        role = (role_map or {}).get(pl, "I")

        events = event_ppms.get(pl, [])

        games_2024 = sum(e["n_games"] for e in events if e["year"] == 2024)
        games_2025 = sum(e["n_games"] for e in events if e["year"] == 2025)
        games_2026 = sum(e["n_games"] for e in events if e["year"] == 2026)
        games_total = games_2024 + games_2025 + games_2026
        n_events = len(events)

        dampened_events, peak_event_ppm, had_outlier = dampen_outlier_events(events)

        # Year weights
        year_weights = get_intl_year_weights(games_2026)
        weighted_ppm = compute_weighted_ppm(dampened_events, year_weights)

        # Bayesian shrinkage
        if games_total == 0:
            weighted_ppm = team_avg_ppm.get(team, global_mean_ppm)
        else:
            shrinkage_games = 6
            wf = games_total / (games_total + shrinkage_games)
            weighted_ppm = wf * weighted_ppm + (1 - wf) * global_mean_ppm

        # Team strength (12%)
        tw = team_wr.get(team, 0.5)
        team_str_adj = (tw - avg_tw) * 0.12 * weighted_ppm

        # Pickrate sentiment
        pick_info = pickrate_dict.get(pl, {})
        avg_pickpct = pick_info.get("avg_pickpct", 0.0)
        pickrate_signal = get_pickrate_signal(avg_pickpct)
        pickrate_weight = max(0.10, 0.40 - games_2026 / 40)
        pickrate_adj = pickrate_signal * pickrate_weight * weighted_ppm

        # Team brand (5%)
        team_pop_val = team_pop.get(team, 0.5)
        brand_adj = (team_pop_val - 0.5) * 0.05 * weighted_ppm

        # Consistency (3%)
        if events and games_total > 2:
            ppms_all = [e["avg_ppm"] for e in events]
            cv = np.std(ppms_all) / (np.mean(ppms_all) + 1e-9)
            consistency_score = max(0, 1 - cv)
        else:
            consistency_score = 0.5
        consistency_adj = (consistency_score - 0.5) * 0.03 * weighted_ppm

        # Small sample dampening
        weighted_ppm_after = weighted_ppm
        if n_events <= 3:
            event_shrink = 0.12 * (4 - n_events) / 3
            weighted_ppm_after = weighted_ppm * (1 - event_shrink) + global_mean_ppm * event_shrink

        PROVEN_EVENTS = 8
        if n_events < PROVEN_EVENTS and weighted_ppm_after > global_mean_ppm * 1.1:
            excess = weighted_ppm_after - global_mean_ppm
            dampen_strength = 0.25 * (PROVEN_EVENTS - n_events) / PROVEN_EVENTS
            weighted_ppm_after -= excess * dampen_strength

        adjusted_ppm = (weighted_ppm_after + team_str_adj +
                       pickrate_adj + brand_adj + consistency_adj)

        records.append({
            "Player": player,
            "Team": team,
            "Role": role,
            "adjusted_ppm": adjusted_ppm,
            "raw_weighted_ppm": weighted_ppm,
            "peak_event_ppm": peak_event_ppm,
            "had_outlier": had_outlier,
            "games_total": games_total,
            "games_2026": games_2026,
            "avg_pickpct": avg_pickpct,
            "uncertainty": get_uncertainty_flag(games_total, games_2026),
        })

    df = pd.DataFrame(records)
    if len(df) == 0:
        return df

    # S-curve mapping
    ranks = df["adjusted_ppm"].rank(pct=True).values
    mapped_vp = np.interp(ranks, target_quantiles, target_vps)
    df["base_vp"] = mapped_vp

    # Potential floor
    ppm_min = df["adjusted_ppm"].min()
    ppm_max = df["adjusted_ppm"].max()
    final_vps = []
    for _, r in df.iterrows():
        base = r["base_vp"]
        if r["had_outlier"] and r["peak_event_ppm"] > 0:
            frac = (r["peak_event_ppm"] - ppm_min) / (ppm_max - ppm_min + 1e-9)
            peak_price = vp_min + frac * (vp_max - vp_min)
            peak_price = np.clip(peak_price, vp_min, vp_max)
            if peak_price > base:
                final_vps.append(base + (peak_price - base) * 0.25)
            else:
                final_vps.append(base)
        else:
            final_vps.append(base)
    df["pre_shift_vp"] = final_vps

    # Role adjustment (faded at extremes)
    base_role_adj = {"D": -0.5, "S": 0.5, "C": 0.3, "I": 0.0}
    role_adjs = []
    for _, r in df.iterrows():
        adj = base_role_adj.get(r["Role"], 0.0)
        vp = r["pre_shift_vp"]
        if vp < vp_min + 1.5:
            fade = max(0, (vp - vp_min) / 1.5)
        elif vp > vp_max - 1.5:
            fade = max(0, (vp_max - vp) / 1.5)
        else:
            fade = 1.0
        role_adjs.append(adj * fade)
    df["pre_shift_vp"] = df["pre_shift_vp"] + role_adjs

    # Shift to target mean
    current_mean = df["pre_shift_vp"].mean()
    shift = target_mean - current_mean
    df["shifted_vp"] = df["pre_shift_vp"] + shift

    # Snap and clip
    df["generated_vp"] = np.round(df["shifted_vp"] * 2) / 2
    df["generated_vp"] = df["generated_vp"].clip(vp_min, vp_max)

    return df


# ============================================================================
#  STEP 2: EVALUATE PRICING
# ============================================================================

def evaluate_pricing(gen_prices, manual_prices, actual_pts, pickrate_dict,
                     event_name, best_n=2):
    """Evaluate generated vs manual pricing accuracy."""
    print(f"\n{'='*70}")
    print(f"  PRICING EVALUATION: {event_name}")
    print(f"{'='*70}")

    # Merge all data
    merged = gen_prices.merge(manual_prices, on="Player", suffixes=("_gen", "_man"))
    if "Team_gen" in merged.columns:
        merged["Team"] = merged["Team_gen"]

    # Add actual performance (best N games)
    merged["actual_best2_pts"] = merged["Player"].map(
        lambda p: compute_best_n_total(actual_pts.get(p, []), best_n))

    # Price accuracy: generated vs manual
    diff_gen = merged["generated_vp"] - merged["manual_vp"]
    diff_man = merged["manual_vp"] - merged["manual_vp"]  # 0 by definition

    print(f"\n  Players compared: {len(merged)}")
    print(f"\n  Generated vs Manual Prices:")
    print(f"    MAE:          {diff_gen.abs().mean():.2f}")
    print(f"    RMSE:         {np.sqrt((diff_gen**2).mean()):.2f}")
    print(f"    Bias:         {diff_gen.mean():+.2f}")
    print(f"    Within 0.5:   {(diff_gen.abs() <= 0.5).sum()}/{len(merged)} "
          f"({100*(diff_gen.abs() <= 0.5).mean():.0f}%)")
    print(f"    Within 1.0:   {(diff_gen.abs() <= 1.0).sum()}/{len(merged)} "
          f"({100*(diff_gen.abs() <= 1.0).mean():.0f}%)")

    # Price-to-performance correlation
    valid = merged[merged["actual_best2_pts"] > 0]
    if len(valid) > 3:
        corr_gen = valid["generated_vp"].corr(valid["actual_best2_pts"])
        corr_man = valid["manual_vp"].corr(valid["actual_best2_pts"])
        # Spearman via rank transform (no scipy needed)
        rank_corr_gen = valid["generated_vp"].rank().corr(valid["actual_best2_pts"].rank())
        rank_corr_man = valid["manual_vp"].rank().corr(valid["actual_best2_pts"].rank())

        print(f"\n  Price-to-Performance Correlation (higher = better pricing):")
        print(f"    {'Metric':<25} {'Generated':>10} {'Manual':>10} {'Winner':>10}")
        print(f"    {'-'*55}")
        pearson_winner = "Generated" if corr_gen > corr_man else "Manual"
        spearman_winner = "Generated" if rank_corr_gen > rank_corr_man else "Manual"
        print(f"    {'Pearson correlation':<25} {corr_gen:>10.3f} {corr_man:>10.3f} {pearson_winner:>10}")
        print(f"    {'Spearman rank corr':<25} {rank_corr_gen:>10.3f} {rank_corr_man:>10.3f} {spearman_winner:>10}")

        # Overall verdict
        gen_score = (1 if corr_gen > corr_man else 0) + (1 if rank_corr_gen > rank_corr_man else 0)
        verdict = "GENERATED wins" if gen_score >= 2 else ("MANUAL wins" if gen_score == 0 else "TIE")
        print(f"\n  >>> VERDICT: {verdict}")
    else:
        print(f"\n  Not enough data for correlation analysis.")
        corr_gen = corr_man = rank_corr_gen = rank_corr_man = 0

    # Pickrate analysis
    merged["pickrate"] = merged["Player"].map(
        lambda p: pickrate_dict.get(p.lower(), {}).get("avg_pickpct", 0))
    picked = merged[merged["pickrate"] > 0]
    if len(picked) > 3:
        pr_corr_gen = picked["generated_vp"].corr(picked["pickrate"])
        pr_corr_man = picked["manual_vp"].corr(picked["pickrate"])
        print(f"\n  Pickrate Correlation:")
        print(f"    Generated vs pickrate: {pr_corr_gen:.3f}")
        print(f"    Manual vs pickrate:    {pr_corr_man:.3f}")

    # Top players table
    top = merged.nlargest(20, "actual_best2_pts")
    print(f"\n  Top 20 Players by Actual Best-2 Points:")
    print(f"    {'Player':<16} {'Team':<18} {'GenVP':>6} {'ManVP':>6} {'Best2':>6} {'Pick%':>6}")
    print(f"    {'-'*60}")
    for _, r in top.iterrows():
        print(f"    {r['Player']:<16} {r['Team']:<18} "
              f"{r['generated_vp']:>6.1f} {r['manual_vp']:>6.1f} "
              f"{r['actual_best2_pts']:>6.1f} {r['pickrate']:>5.1f}%")

    return {
        "merged": merged,
        "corr_gen": corr_gen if len(valid) > 3 else None,
        "corr_man": corr_man if len(valid) > 3 else None,
    }


# ============================================================================
#  STEP 3: BUILD KICKOFF TEAM
# ============================================================================

def estimate_expected_best_n(player_games, n_best, n_expected_games,
                             n_simulations=5000):
    """Bootstrap estimate of expected best-N-game score from historical data."""
    if not player_games:
        return 0.0
    pts_history = [g["pts"] for g in player_games]
    if len(pts_history) == 0:
        return 0.0

    total = 0.0
    for _ in range(n_simulations):
        sampled = RNG.choice(pts_history, size=n_expected_games, replace=True)
        top_n = np.sort(sampled)[-n_best:]
        total += top_n.sum()
    return total / n_simulations


def estimate_ceiling_best_n(player_games, n_best, n_expected_games,
                            n_simulations=5000, percentile=90):
    """Bootstrap estimate of ceiling (e.g. 90th percentile) best-N score.
    Used for IGL selection — we want the player most likely to pop off."""
    if not player_games:
        return 0.0
    pts_history = [g["pts"] for g in player_games]
    if len(pts_history) == 0:
        return 0.0

    scores = []
    for _ in range(n_simulations):
        sampled = RNG.choice(pts_history, size=n_expected_games, replace=True)
        top_n = np.sort(sampled)[-n_best:]
        scores.append(top_n.sum())
    return float(np.percentile(scores, percentile))


def estimate_team_games(team_wr, bracket_type="kickoff"):
    """Estimate expected games based on team win rate."""
    wr = team_wr
    if bracket_type == "kickoff":
        # Double elim: min 3 games (lose twice fast), max ~8 for finalists
        if wr > 0.6:
            return 7
        elif wr > 0.45:
            return 5
        else:
            return 3
    return 3


def compute_pool_expected_pts(player_pool, historical_games, team_wr,
                               n_best, bracket_type="kickoff",
                               fixed_n_games=None, role_map=None):
    """Compute expected points for ALL players in a pool.

    Args:
        player_pool: DataFrame with Player, Team columns
        historical_games: dict player_lower -> list of game dicts
        team_wr: dict team -> win rate
        n_best: how many best games to count (None = all games count)
        bracket_type: "kickoff" for bracket estimation of games
        fixed_n_games: if set, every player plays this many games (e.g. 5 for Stage 1)
        role_map: dict player_lower -> role letter

    Returns:
        dict: player_name -> expected_pts
    """
    result = {}
    for _, row in player_pool.iterrows():
        player = row["Player"]
        team = row["Team"]
        pl = player.lower()

        hist = historical_games.get(pl, [])
        if not hist:
            result[player] = 0.0
            continue

        if fixed_n_games is not None:
            n_games = fixed_n_games
        else:
            wr = team_wr.get(team, 0.5)
            n_games = estimate_team_games(wr, bracket_type)

        if n_best is not None:
            expected = estimate_expected_best_n(hist, n_best, n_games)
        else:
            # All games count — bootstrap n_games, sum all
            pts_history = [g["pts"] for g in hist]
            total = 0.0
            n_sim = 5000
            for _ in range(n_sim):
                sampled = RNG.choice(pts_history, size=n_games, replace=True)
                total += sampled.sum()
            expected = total / n_sim

        result[player] = expected
    return result


def build_kickoff_team(player_pool_with_prices, player_historical_games,
                       team_wr, role_map, budget=100, n_iter=15000,
                       pickrate_dict=None):
    """Build optimal 11-player team for Kickoff.

    Uses value-ratio (expected/price) scoring — the correct knapsack heuristic.
    Runs multiple random seeds and keeps the best team found across all seeds.
    """
    # Estimate expected points for each player
    players = []
    for _, row in player_pool_with_prices.iterrows():
        player = row["Player"]
        team = row["Team"]
        pl = player.lower()
        price = row["manual_vp"]
        role = role_map.get(pl, "I")

        wr = team_wr.get(team, 0.5)
        n_games = estimate_team_games(wr)
        hist_games = player_historical_games.get(pl, [])

        expected = estimate_expected_best_n(hist_games, KICKOFF_BEST_N, n_games)
        ceiling = estimate_ceiling_best_n(hist_games, KICKOFF_BEST_N, n_games)

        # Skip China players
        if team in CHINA_TEAMS_INTL:
            continue

        players.append({
            "Player": player,
            "Team": team,
            "price": price,
            "role": role,
            "expected_best2": expected,
            "ceiling_best3": ceiling,
            "est_games": n_games,
        })

    # Run optimizer with multiple random seeds to explore diverse compositions
    best_team = None
    best_score = -1

    for seed_offset in range(5):
        rng = np.random.default_rng(2026 + seed_offset * 1000)

        for iteration in range(n_iter):
            noise = rng.normal(0, 0.5, len(players))
            scored = [(p, p["expected_best2"] / (p["price"] + 0.1) + noise[i])
                      for i, p in enumerate(players)]
            scored.sort(key=lambda x: x[1], reverse=True)

            team = []
            team_vp = 0.0
            team_counts = {}
            role_counts = {"D": 0, "C": 0, "I": 0, "S": 0}
            wc_used = 0

            for p, _ in scored:
                if len(team) >= KICKOFF_SQUAD:
                    break
                vp = p["price"]
                t = p["Team"]
                role = p["role"]

                if team_vp + vp > budget:
                    continue
                if team_counts.get(t, 0) >= 2:
                    continue
                remaining = KICKOFF_SQUAD - len(team) - 1
                if remaining > 0 and (budget - team_vp - vp) < remaining * INTL_VP_MIN:
                    continue

                if role_counts.get(role, 0) < KICKOFF_ROLE_SLOTS.get(role, 0):
                    role_counts[role] += 1
                elif wc_used < KICKOFF_WC_SLOTS:
                    wc_used += 1
                else:
                    continue

                team.append(p)
                team_vp += vp
                team_counts[t] = team_counts.get(t, 0) + 1

            if len(team) == KICKOFF_SQUAD:
                rc = {"D": 0, "C": 0, "I": 0, "S": 0}
                for p in team:
                    r = p["role"]
                    if r in rc:
                        rc[r] += 1
                if all(rc[r] >= KICKOFF_ROLE_SLOTS[r] for r in KICKOFF_ROLE_SLOTS):
                    total_pts = sum(p["expected_best2"] for p in team)
                    # Try top 3 IGL candidates (highest ceiling)
                    candidates = sorted(team, key=lambda p: p.get("ceiling_best3",
                                        p["expected_best2"]), reverse=True)[:3]
                    for igl_cand in candidates:
                        score = total_pts + igl_cand["expected_best2"]
                        if score > best_score:
                            best_score = score
                            best_team = {
                                "players": list(team),
                                "total_vp": round(team_vp, 2),
                                "expected_pts": round(total_pts, 1),
                                "expected_pts_with_igl": round(score, 1),
                                "igl": igl_cand["Player"],
                            }

    return best_team


def score_team_actual(team, actual_pts, n_best=2, igl_player=None):
    """Score a team against actual results."""
    total = 0.0
    player_scores = []
    for p in team["players"]:
        games = actual_pts.get(p["Player"], [])
        best_n = compute_best_n_total(games, n_best)
        is_igl = p["Player"] == (igl_player or team.get("igl"))
        multiplier = 2.0 if is_igl else 1.0
        player_scores.append({
            "Player": p["Player"],
            "Team": p["Team"],
            "price": p["price"],
            "role": p["role"],
            "best_n_pts": best_n,
            "igl_mult": multiplier,
            "total_pts": best_n * multiplier,
            "games_played": len(games),
        })
        total += best_n * multiplier
    return total, player_scores


# ============================================================================
#  STEP 4: BUILD SANTIAGO TEAMS
# ============================================================================

def build_santiago_team(player_pool_with_prices, player_expected_pts,
                        role_map, budget=50, n_iter=8000, max_china=1):
    """Build optimal 6-player Santiago team.

    Each player is assigned a 'slot': their role letter (D/C/I/S) for mandatory
    role slots, or 'W' for wildcard slots. Transfers must respect slots.
    China players are limited to max_china (default 1) as budget fillers only.
    """
    players = []
    for _, row in player_pool_with_prices.iterrows():
        player = row["Player"]
        team = row["Team"]
        pl = player.lower()
        price = row["manual_vp"]
        role = role_map.get(pl, "I")
        exp = player_expected_pts.get(player, 0)

        # China players: heavily penalized but included as budget fillers
        is_china = team in CHINA_TEAMS_INTL
        if is_china:
            exp = 0.1

        players.append({
            "Player": player,
            "Team": team,
            "price": price,
            "role": role,
            "expected_best2": exp,
            "is_china": is_china,
        })

    if not players:
        return None

    # Compute actual minimum price in pool for budget floor check
    pool_min_price = min(p["price"] for p in players)

    best_team = None
    best_score = -1

    for iter_n in range(n_iter):
        # Phase 1: Fill mandatory role slots first (1D, 1C, 1I, 1S)
        # Phase 2: Fill wildcards with best remaining
        team = []
        team_vp = 0.0
        team_counts = {}
        used_players = set()
        china_count = 0
        slot_assignments = {}  # Player -> slot ("D"/"C"/"I"/"S"/"W")

        # Phase 1: pick one player per mandatory role
        role_order = list(["D", "C", "I", "S"])
        if RNG.random() > 0.3:
            RNG.shuffle(role_order)

        for role in role_order:
            candidates = [p for p in players
                         if p["role"] == role and p["Player"] not in used_players
                         and team_counts.get(p["Team"], 0) < 2
                         and not (p.get("is_china") and china_count >= max_china)]
            if not candidates:
                break
            scored_c = [(p, p["expected_best2"] / (p["price"] + 0.1)
                        + float(RNG.normal(0, 0.3)))
                       for p in candidates]
            scored_c.sort(key=lambda x: x[1], reverse=True)

            picked = False
            for p, _ in scored_c:
                remaining_after = SANTIAGO_SQUAD - len(team) - 1
                budget_left = budget - team_vp - p["price"]
                if budget_left < 0:
                    continue
                if remaining_after > 0 and budget_left < remaining_after * pool_min_price:
                    continue
                team.append(p)
                team_vp += p["price"]
                team_counts[p["Team"]] = team_counts.get(p["Team"], 0) + 1
                used_players.add(p["Player"])
                slot_assignments[p["Player"]] = role  # assigned to role slot
                if p.get("is_china"):
                    china_count += 1
                picked = True
                break
            if not picked:
                break

        if len(team) < 4:  # couldn't fill all 4 mandatory roles
            continue

        # Phase 2: fill remaining wildcard slots
        wc_candidates = [p for p in players
                        if p["Player"] not in used_players
                        and team_counts.get(p["Team"], 0) < 2
                        and not (p.get("is_china") and china_count >= max_china)]
        scored_wc = [(p, p["expected_best2"] / (p["price"] + 0.1)
                     + float(RNG.normal(0, 0.3)))
                    for p in wc_candidates]
        scored_wc.sort(key=lambda x: x[1], reverse=True)

        for p, _ in scored_wc:
            if len(team) >= SANTIAGO_SQUAD:
                break
            if team_vp + p["price"] > budget:
                continue
            if team_counts.get(p["Team"], 0) >= 2:
                continue
            if p.get("is_china") and china_count >= max_china:
                continue
            remaining = SANTIAGO_SQUAD - len(team) - 1
            if remaining > 0 and (budget - team_vp - p["price"]) < remaining * pool_min_price:
                continue
            team.append(p)
            team_vp += p["price"]
            team_counts[p["Team"]] = team_counts.get(p["Team"], 0) + 1
            used_players.add(p["Player"])
            slot_assignments[p["Player"]] = "W"  # wildcard slot
            if p.get("is_china"):
                china_count += 1

        if len(team) == SANTIAGO_SQUAD:
            # Tag each player dict with their slot
            for p in team:
                p["slot"] = slot_assignments[p["Player"]]

            total_pts = sum(p["expected_best2"] for p in team)
            igl = max(team, key=lambda p: p["expected_best2"])
            total_with_igl = total_pts + igl["expected_best2"]

            if total_with_igl > best_score:
                best_score = total_with_igl
                best_team = {
                    "players": [dict(p) for p in team],
                    "total_vp": round(team_vp, 2),
                    "expected_pts": round(total_pts, 1),
                    "expected_pts_with_igl": round(total_with_igl, 1),
                    "igl": igl["Player"],
                }

    return best_team


def santiago_transfers(current_team, available_players, player_expected_pts,
                       role_map, budget=50, max_transfers=2, n_iter=8000,
                       gw_teams=None):
    """Optimize transfers for Santiago GW2/GW3.

    Key rules:
    - Transfers must respect slot assignments: a role-slot player (D/C/I/S)
      can only be replaced by a player of the SAME role. A wildcard-slot player
      (W) can be replaced by any role.
    - Eliminated players (team not in current GW's team list) MUST be transferred out.
    - China players are excluded as replacement candidates.
    - A player transferred IN cannot be transferred OUT in the same GW.

    gw_teams: set of team names playing in this GW (for eliminated detection)
    """
    current_names = {p["Player"] for p in current_team["players"]}
    current_vp = current_team["total_vp"]

    # Build candidate list (exclude current team and China)
    all_candidates = []
    for _, row in available_players.iterrows():
        player = row["Player"]
        if player in current_names:
            continue
        team = row["Team"]
        if team in CHINA_TEAMS_INTL:
            continue
        pl = player.lower()
        price = row["manual_vp"]
        role = role_map.get(pl, "I")
        exp = player_expected_pts.get(player, 0)
        all_candidates.append({
            "Player": player,
            "Team": team,
            "price": price,
            "role": role,
            "expected_best2": exp,
        })

    # Update current player expectations
    for p in current_team["players"]:
        p["expected_best2"] = player_expected_pts.get(p["Player"], 0)

    # Identify eliminated players: team not in this GW's active teams
    # OR expected 0 pts (covers China players from previous GW)
    eliminated_players = set()
    for p in current_team["players"]:
        if gw_teams and p["Team"] not in gw_teams:
            eliminated_players.add(p["Player"])
        elif p["Team"] in CHINA_TEAMS_INTL:
            eliminated_players.add(p["Player"])
        elif player_expected_pts.get(p["Player"], 0) == 0:
            eliminated_players.add(p["Player"])

    print(f"    Eliminated: {eliminated_players or 'none'}")

    best_result = None
    best_score = -1

    for _ in range(n_iter):
        squad = [dict(p) for p in current_team["players"]]
        squad_vp = current_vp
        transfers_made = []
        already_transferred_out = set()
        transferred_in = set()  # prevent transfer-in-then-out

        # Force transfers for eliminated players first, then consider optional upgrades
        n_forced = min(len(eliminated_players), max_transfers)
        n_extra = int(RNG.integers(0, max(0, max_transfers - n_forced) + 1))
        target_transfers = n_forced + n_extra
        transfers_done = 0

        for attempt in range(target_transfers + 5):
            if transfers_done >= min(target_transfers, max_transfers):
                break

            # Pick who to transfer out (never transfer out someone we just brought in)
            remaining_elim = [p for p in squad
                             if p["Player"] in eliminated_players
                             and p["Player"] not in already_transferred_out
                             and p["Player"] not in transferred_in]
            if remaining_elim:
                remaining_elim.sort(key=lambda p: (0 if p.get("slot") == "W" else 1,
                                                    p["expected_best2"]))
                out_player = remaining_elim[0]
            elif transfers_done < target_transfers:
                non_elim = [p for p in squad
                           if p["Player"] not in already_transferred_out
                           and p["Player"] not in eliminated_players
                           and p["Player"] not in transferred_in]
                if not non_elim:
                    break
                non_elim_scored = [(p, p["expected_best2"] + float(RNG.normal(0, 1.0)))
                                   for p in non_elim]
                non_elim_scored.sort(key=lambda x: x[1])
                out_player = non_elim_scored[0][0]
            else:
                break

            out_slot = out_player.get("slot", "W")

            # Find valid replacements respecting slot constraints
            candidates = []
            squad_names = {p["Player"] for p in squad}
            for c in all_candidates:
                if c["Player"] in squad_names:
                    continue
                if c["price"] > budget - squad_vp + out_player["price"]:
                    continue
                # Slot constraint: role slot requires same role, wildcard allows any
                if out_slot != "W" and c["role"] != out_slot:
                    continue
                # Team limit
                tc = sum(1 for p in squad if p["Team"] == c["Team"]
                        and p["Player"] != out_player["Player"])
                if tc >= 2:
                    continue
                candidates.append(c)

            if not candidates:
                already_transferred_out.add(out_player["Player"])
                continue

            noise = RNG.normal(0, 0.8, len(candidates))
            scored_cands = [(c, c["expected_best2"] + float(noise[i]))
                           for i, c in enumerate(candidates)]
            scored_cands.sort(key=lambda x: x[1], reverse=True)

            swapped = False
            for cand, _ in scored_cands[:5]:
                new_vp = squad_vp - out_player["price"] + cand["price"]
                if new_vp > budget:
                    continue
                cand_copy = dict(cand)
                cand_copy["slot"] = out_slot
                squad = [p for p in squad if p["Player"] != out_player["Player"]]
                squad.append(cand_copy)
                squad_vp = new_vp
                transfers_made.append((out_player["Player"], cand["Player"]))
                already_transferred_out.add(out_player["Player"])
                transferred_in.add(cand["Player"])
                transfers_done += 1
                swapped = True
                break

            if not swapped:
                already_transferred_out.add(out_player["Player"])

        total_pts = sum(p["expected_best2"] for p in squad)
        if squad:
            igl = max(squad, key=lambda p: p["expected_best2"])
            total_with_igl = total_pts + igl["expected_best2"]
        else:
            total_with_igl = 0
            igl = {"Player": "N/A"}

        if total_with_igl > best_score:
            best_score = total_with_igl
            best_result = {
                "players": [dict(p) for p in squad],
                "total_vp": round(squad_vp, 2),
                "expected_pts": round(total_pts, 1),
                "expected_pts_with_igl": round(total_with_igl, 1),
                "igl": igl["Player"],
                "transfers": list(transfers_made),
            }

    return best_result


def compute_santiago_expected_pts(player_pool, historical_games, team_wr,
                                  n_games=3, n_best=2,
                                  tournament_games=None, tournament_weight=0.0,
                                  avoid_china=True):
    """Estimate expected best-2-of-N for Santiago players.

    Heavily factors in team strength: stronger teams advance further,
    play more games, and are more likely to be available in future GWs.

    tournament_games: dict of player -> list of game dicts from prior Santiago GWs
    tournament_weight: how much to weight tournament form (0-1)
    team_wr: dict team -> win rate (0-1)
    """
    result = {}
    avg_wr = np.mean(list(team_wr.values())) if team_wr else 0.5

    for _, row in player_pool.iterrows():
        player = row["Player"]
        team = row["Team"]
        pl = player.lower()

        # Skip China players
        if avoid_china and team in CHINA_TEAMS_INTL:
            result[player] = 0.0
            continue

        hist = historical_games.get(pl, [])

        # Historical estimate
        hist_expected = estimate_expected_best_n(hist, n_best, n_games, 3000)

        # Tournament form
        if tournament_games and player in tournament_games and tournament_weight > 0:
            tourn_pts = [g["pts"] for g in tournament_games[player]]
            if tourn_pts:
                tourn_avg = np.mean(tourn_pts)
                tourn_expected = tourn_avg * min(n_games, len(tourn_pts)) * 0.85
                tourn_expected = max(tourn_expected, sum(sorted(tourn_pts, reverse=True)[:n_best]))
                hist_expected = (1 - tournament_weight) * hist_expected + tournament_weight * tourn_expected

        # Team strength multiplier: strong teams advance, weak teams go home
        # Scale: wr=0.6+ -> 1.25x, wr=0.5 -> 1.0x, wr=0.3 -> 0.7x
        wr = team_wr.get(team, 0.5)
        team_mult = 1.0 + (wr - avg_wr) * 1.5
        team_mult = max(0.6, min(1.4, team_mult))

        result[player] = hist_expected * team_mult
    return result


# ============================================================================
#  EXCEL OUTPUT
# ============================================================================

def create_intl_excel(event_name, gen_prices_df, manual_prices_df,
                      actual_pts, team_data, eval_results,
                      per_gw_pts=None, expected_pts_dict=None):
    """Create Excel output for an international event.

    per_gw_pts: optional list of (label, pts_dict) for per-GW columns in Prices sheet.
                e.g. [("GW1", gw1_pts), ("GW2", gw2_pts), ("GW3", gw3_pts)]
                If None, uses single column from actual_pts.
    expected_pts_dict: optional dict player -> expected total pts for "Expected Pts" column.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ---- Sheet 1: Prices ----
    ws1 = wb.active
    ws1.title = "Prices"

    best_n = eval_results.get("best_n", 2)
    header_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")

    # Build headers dynamically based on per_gw_pts
    base_headers = ["Player", "Team", "Role", "Manual VP", "Generated VP", "Diff"]
    exp_headers = ["Expected Pts"] if expected_pts_dict else []
    if per_gw_pts:
        pts_headers = [f"Best-{best_n} {label}" for label, _ in per_gw_pts]
    else:
        pts_headers = [f"Actual Best-{best_n} Pts"]
    end_headers = ["Pick%", "Uncertainty"]
    headers = base_headers + exp_headers + pts_headers + end_headers

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border

    # Merge gen and manual prices
    merged = gen_prices_df[["Player", "Team", "Role", "generated_vp", "uncertainty"]].merge(
        manual_prices_df[["Player", "manual_vp"]], on="Player", how="left"
    )

    # Compute per-GW or single actual points columns
    if per_gw_pts:
        for label, gw_pts in per_gw_pts:
            col_name = f"pts_{label}"
            merged[col_name] = merged["Player"].map(
                lambda p, gp=gw_pts: compute_best_n_total(gp.get(p, []), best_n))
        pts_col_names = [f"pts_{label}" for label, _ in per_gw_pts]
    else:
        merged["pts_all"] = merged["Player"].map(
            lambda p: compute_best_n_total(actual_pts.get(p, []), best_n))
        pts_col_names = ["pts_all"]

    if expected_pts_dict:
        merged["expected_pts"] = merged["Player"].map(
            lambda p: expected_pts_dict.get(p, 0.0))

    pr_dict = eval_results.get("pickrate_dict", {})
    merged["pickrate"] = merged["Player"].map(
        lambda p: pr_dict.get(p.lower(), {}).get("avg_pickpct", 0))
    merged["diff"] = merged["generated_vp"] - merged["manual_vp"]
    merged = merged.sort_values("generated_vp", ascending=False)

    light_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    light_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    for i, (_, row) in enumerate(merged.iterrows(), 2):
        col = 1
        ws1.cell(row=i, column=col, value=row["Player"]).border = thin_border; col += 1
        ws1.cell(row=i, column=col, value=row["Team"]).border = thin_border; col += 1
        ws1.cell(row=i, column=col, value=row["Role"]).border = thin_border; col += 1
        ws1.cell(row=i, column=col, value=row.get("manual_vp", "")).border = thin_border; col += 1
        ws1.cell(row=i, column=col, value=row["generated_vp"]).border = thin_border; col += 1
        c_diff = ws1.cell(row=i, column=col, value=round(row["diff"], 1) if pd.notna(row["diff"]) else "")
        c_diff.border = thin_border
        if pd.notna(row["diff"]):
            if abs(row["diff"]) > 2:
                c_diff.fill = light_red
            elif abs(row["diff"]) <= 1:
                c_diff.fill = light_green
        col += 1
        # Expected pts column (if provided)
        if expected_pts_dict:
            ws1.cell(row=i, column=col,
                     value=round(row.get("expected_pts", 0), 1)).border = thin_border
            col += 1
        # Points columns (1 or 3)
        for pc in pts_col_names:
            ws1.cell(row=i, column=col, value=round(row[pc], 1)).border = thin_border
            col += 1
        ws1.cell(row=i, column=col, value=round(row["pickrate"], 1)).border = thin_border; col += 1
        ws1.cell(row=i, column=col, value=row["uncertainty"] if pd.notna(row["uncertainty"]) else "").border = thin_border

    from openpyxl.utils import get_column_letter
    base_widths = [18, 20, 6, 10, 12, 8]
    exp_widths = [14] if expected_pts_dict else []
    pts_widths = [14] * len(pts_col_names)
    end_widths = [8, 30]
    widths = base_widths + exp_widths + pts_widths + end_widths
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 2: Teams ----
    ws2 = wb.create_sheet("Teams")

    team_header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")

    row_num = 1
    for td in team_data:
        label = td["label"]
        cell = ws2.cell(row=row_num, column=1, value=label)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = team_header_fill
        for c in range(2, 10):
            ws2.cell(row=row_num, column=c).fill = team_header_fill
        row_num += 1

        # Column headers
        gw_best_n = td.get("best_n", 2)
        cols = ["Player", "Team", "Slot", "Role", "VP", f"Expected Best-{gw_best_n}",
                f"Actual Best-{gw_best_n}", "IGL", "Transfers"]
        subheader_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws2.cell(row=row_num, column=col_idx, value=col_name)
            cell.font = Font(bold=True, size=10)
            cell.fill = subheader_fill
            cell.border = thin_border
        row_num += 1

        team_info = td["team"]
        if team_info is None:
            ws2.cell(row=row_num, column=1, value="FAILED to build team")
            row_num += 2
            continue

        # Transfers
        transfers_str = ""
        if team_info.get("transfers"):
            parts = [f"{t[0]} -> {t[1]}" for t in team_info["transfers"]]
            transfers_str = " | ".join(parts)

        igl_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

        # Use per-GW actual_pts if provided, otherwise fall back to event-wide
        gw_actual = td.get("actual_pts", actual_pts)

        for p in team_info["players"]:
            games = gw_actual.get(p["Player"], [])
            actual_best2 = compute_best_n_total(games, 2)
            is_igl = p["Player"] == team_info.get("igl")

            ws2.cell(row=row_num, column=1, value=p["Player"]).border = thin_border
            ws2.cell(row=row_num, column=2, value=p["Team"]).border = thin_border
            ws2.cell(row=row_num, column=3, value=p.get("slot", "?")).border = thin_border
            ws2.cell(row=row_num, column=4, value=p["role"]).border = thin_border
            ws2.cell(row=row_num, column=5, value=p["price"]).border = thin_border
            ws2.cell(row=row_num, column=6, value=round(p.get("expected_best2", 0), 1)).border = thin_border
            ws2.cell(row=row_num, column=7, value=round(actual_best2, 1)).border = thin_border
            igl_cell = ws2.cell(row=row_num, column=8, value="IGL" if is_igl else "")
            igl_cell.border = thin_border
            if is_igl:
                igl_cell.font = Font(bold=True, color="FF0000")
                igl_cell.fill = igl_fill
            row_num += 1

        # Totals row
        actual_total, _ = score_team_actual(team_info, gw_actual, 2)
        ws2.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
        ws2.cell(row=row_num, column=5, value=team_info["total_vp"]).font = Font(bold=True)
        ws2.cell(row=row_num, column=6, value=team_info["expected_pts_with_igl"]).font = Font(bold=True)
        ws2.cell(row=row_num, column=7, value=round(actual_total, 1)).font = Font(bold=True, color="0000FF")
        if transfers_str:
            ws2.cell(row=row_num, column=9, value=transfers_str).font = Font(italic=True, size=9)
        row_num += 2

    widths2 = [18, 20, 6, 6, 7, 14, 14, 6, 40]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    output_path = os.path.join(DIR, f"VFL_2026_{event_name}_Pricing.xlsx")
    wb.save(output_path)
    print(f"\n  Excel saved to: {output_path}")
    return output_path


# ============================================================================
#  MAIN
# ============================================================================

def main():
    print("=" * 70)
    print("  VFL 2026 INTERNATIONAL EVENTS — Pricing & Team Building")
    print("=" * 70)

    # Load full 2026 CSV for actual results and manual prices
    csv_2026 = load_csv(2026)
    # For international events, include China teams
    csv_2026_full = csv_2026.copy()  # includes China

    # Load pickrate data
    pickrate_df = load_pickrate_data()
    all_pickrate_dict = compute_player_pickrates(pickrate_df) if pickrate_df is not None else {}

    # ================================================================
    #  STEP 1A: KICKOFF 2026 PRICES (train on 2024+2025 only)
    # ================================================================
    print("\n" + "=" * 70)
    print("  STEP 1A: GENERATE KICKOFF 2026 PRICES")
    print("=" * 70)

    # Load training data (2024+2025 only, include China for international)
    kickoff_train = load_data_up_to(2025, include_china=True)
    print(f"  Training data: {len(kickoff_train)} rows")

    # Player pool and manual prices from dedicated CSV
    kickoff_manual = load_manual_intl_prices("kickoff")
    kickoff_players = kickoff_manual[["Player", "Team"]].copy()
    print(f"  Kickoff player pool: {len(kickoff_players)} players")
    print(f"  Kickoff manual prices: {len(kickoff_manual)} players, "
          f"range=[{kickoff_manual['manual_vp'].min():.1f}, {kickoff_manual['manual_vp'].max():.1f}]")

    # Team win rates from training data
    played_train = kickoff_train[kickoff_train["P?"] == 1]
    kickoff_team_wr = compute_team_win_rates(played_train)
    kickoff_team_pop = compute_team_popularity(pickrate_df, kickoff_train) if pickrate_df is not None else {}

    # Pickrate: only use events 1-3 (pre-Kickoff)
    kickoff_pickrate = {}
    if pickrate_df is not None:
        # Filter to events before Kickoff 2026 (events_appeared from pre-2026)
        # The pickrate_summary doesn't have per-event breakdown, so use all
        # available data (it's aggregated). This is approximate.
        kickoff_pickrate = all_pickrate_dict

    # Load Santiago prices early for role map
    santiago_manual = load_manual_intl_prices("santiago")

    # Build role map from both price files
    # Santiago has full names (Duelist->D, Controller->C, Initiator->I, Sentinel->S)
    role_name_map = {"Duelist": "D", "Controller": "C", "Initiator": "I", "Sentinel": "S",
                     "D": "D", "C": "C", "I": "I", "S": "S"}
    role_map = {}
    for _, r in kickoff_manual.iterrows():
        role_map[r["Player"].lower()] = role_name_map.get(r.get("Role", "I"), "I")
    for _, r in santiago_manual.iterrows():
        role_map[r["Player"].lower()] = role_name_map.get(r.get("Role", "I"), "I")

    # Generate prices
    kickoff_gen = generate_intl_prices(
        kickoff_train, kickoff_players, kickoff_pickrate, kickoff_team_pop,
        kickoff_team_wr, target_mean=KICKOFF_BUDGET / KICKOFF_SQUAD,
        vp_min=INTL_VP_MIN, vp_max=INTL_VP_MAX, role_map=role_map,
    )
    print(f"  Generated prices: {len(kickoff_gen)} players")
    print(f"  Mean: {kickoff_gen['generated_vp'].mean():.2f}, "
          f"Median: {kickoff_gen['generated_vp'].median():.2f}")

    # ================================================================
    #  STEP 1B: SANTIAGO 2026 PRICES (train on 2024+2025+Kickoff)
    # ================================================================
    print("\n" + "=" * 70)
    print("  STEP 1B: GENERATE SANTIAGO 2026 PRICES")
    print("=" * 70)

    santiago_train = load_data_up_to(2026, max_stage="Kickoff", include_china=True)
    print(f"  Training data: {len(santiago_train)} rows")

    santiago_players = santiago_manual[["Player", "Team"]].copy()
    print(f"  Santiago player pool: {len(santiago_players)} players")
    print(f"  Santiago manual prices: {len(santiago_manual)} players, "
          f"range=[{santiago_manual['manual_vp'].min():.1f}, {santiago_manual['manual_vp'].max():.1f}]")

    played_santiago = santiago_train[santiago_train["P?"] == 1]
    santiago_team_wr = compute_team_win_rates(played_santiago)
    santiago_team_pop = compute_team_popularity(pickrate_df, santiago_train) if pickrate_df is not None else {}

    santiago_gen = generate_intl_prices(
        santiago_train, santiago_players, all_pickrate_dict, santiago_team_pop,
        santiago_team_wr, target_mean=SANTIAGO_BUDGET / SANTIAGO_SQUAD,
        vp_min=INTL_VP_MIN, vp_max=INTL_VP_MAX, role_map=role_map,
    )
    print(f"  Generated prices: {len(santiago_gen)} players")
    print(f"  Mean: {santiago_gen['generated_vp'].mean():.2f}, "
          f"Median: {santiago_gen['generated_vp'].median():.2f}")

    # ================================================================
    #  STEP 2: EVALUATE PRICING
    # ================================================================
    print("\n" + "=" * 70)
    print("  STEP 2: EVALUATE PRICING ACCURACY")
    print("=" * 70)

    kickoff_actual_pts = get_player_actual_game_pts(csv_2026_full, "Kickoff")
    santiago_actual_pts = get_player_actual_game_pts(csv_2026_full, "Santiago")

    kickoff_eval = evaluate_pricing(
        kickoff_gen, kickoff_manual, kickoff_actual_pts,
        all_pickrate_dict, "KICKOFF 2026", best_n=KICKOFF_BEST_N
    )
    kickoff_eval["pickrate_dict"] = all_pickrate_dict
    kickoff_eval["best_n"] = KICKOFF_BEST_N

    santiago_eval = evaluate_pricing(
        santiago_gen, santiago_manual, santiago_actual_pts,
        all_pickrate_dict, "SANTIAGO 2026"
    )
    santiago_eval["pickrate_dict"] = all_pickrate_dict
    santiago_eval["best_n"] = SANTIAGO_BEST_N

    # ================================================================
    #  STEP 3: BUILD KICKOFF TEAM
    # ================================================================
    print("\n" + "=" * 70)
    print("  STEP 3: BUILD KICKOFF 2026 TEAM")
    print("=" * 70)

    # Collect historical per-game data for bootstrap
    kickoff_hist_train = kickoff_train[kickoff_train["P?"] == 1]
    player_historical_games = {}
    for (player,), grp in kickoff_hist_train.groupby(["Player"]):
        pl = player.lower()
        games = []
        for _, r in grp.iterrows():
            games.append({"pts": r["Pts"], "ppm": r["PPM"]})
        player_historical_games[pl] = games

    print(f"  Historical game data: {len(player_historical_games)} players")
    print(f"  Building team (11 players, 100 VP, best 3 games)...")

    kickoff_team = build_kickoff_team(
        kickoff_manual, player_historical_games, kickoff_team_wr,
        role_map, budget=KICKOFF_BUDGET, n_iter=15000,
        pickrate_dict=all_pickrate_dict
    )

    if kickoff_team:
        print(f"\n  KICKOFF TEAM:")
        print(f"    {'Player':<16} {'Team':<18} {'Role':>4} {'VP':>6} {'ExpB2':>7}")
        print(f"    {'-'*55}")
        for p in sorted(kickoff_team["players"], key=lambda x: x["expected_best2"], reverse=True):
            igl = " *IGL*" if p["Player"] == kickoff_team["igl"] else ""
            print(f"    {p['Player']:<16} {p['Team']:<18} {p['role']:>4} "
                  f"{p['price']:>6.1f} {p['expected_best2']:>7.1f}{igl}")
        print(f"    Total VP: {kickoff_team['total_vp']:.1f}  |  "
              f"Expected (with IGL): {kickoff_team['expected_pts_with_igl']:.1f}")

        # Score against actual
        actual_total, player_scores = score_team_actual(
            kickoff_team, kickoff_actual_pts, KICKOFF_BEST_N)

        print(f"\n  ACTUAL RESULTS:")
        print(f"    {'Player':<16} {'Team':<18} {'Best2':>6} {'Mult':>5} {'Total':>7}")
        print(f"    {'-'*55}")
        for ps in sorted(player_scores, key=lambda x: x["total_pts"], reverse=True):
            igl = "2x" if ps["igl_mult"] > 1 else "1x"
            print(f"    {ps['Player']:<16} {ps['Team']:<18} {ps['best_n_pts']:>6.1f} "
                  f"{igl:>5} {ps['total_pts']:>7.1f}")
        print(f"    TOTAL ACTUAL POINTS: {actual_total:.1f}")
    else:
        print("  FAILED to build a valid team!")

    # ================================================================
    #  STEP 4: BUILD SANTIAGO TEAMS (3 GWs)
    # ================================================================
    print("\n" + "=" * 70)
    print("  STEP 4: BUILD SANTIAGO 2026 TEAMS")
    print("=" * 70)

    # Historical data up to Kickoff for initial Santiago team building
    santiago_hist_train = santiago_train[santiago_train["P?"] == 1]
    santiago_hist_games = {}
    for (player,), grp in santiago_hist_train.groupby(["Player"]):
        pl = player.lower()
        games = []
        for _, r in grp.iterrows():
            games.append({"pts": r["Pts"], "ppm": r["PPM"]})
        santiago_hist_games[pl] = games

    # --- GW1: Swiss Stage ---
    print("\n  --- GW1: Swiss Stage ---")
    print(f"  Swiss teams: {', '.join(SANTIAGO_SWISS_TEAMS)}")

    # Filter to Swiss teams only
    swiss_players = santiago_manual[
        santiago_manual["Team"].isin(SANTIAGO_SWISS_TEAMS)
    ].copy()
    print(f"  Swiss player pool: {len(swiss_players)} players")

    # Estimate expected best-2-of-3 for Swiss (3 rounds)
    # CRITICAL: For GW1, heavily boost players from strong teams that will advance
    # to GW2/GW3. Picking a weak team player means 0 pts in future GWs and wasted
    # transfer slots. We add a "tournament longevity" multiplier on top.
    gw1_expected = compute_santiago_expected_pts(
        swiss_players, santiago_hist_games, santiago_team_wr,
        n_games=3, n_best=2
    )
    # Apply tournament longevity bonus: players from strong teams will score
    # in GW2+GW3 too, while weak team players become dead weight eating transfers.
    # This is THE most important factor for GW1 picks.
    avg_wr = np.mean(list(santiago_team_wr.values())) if santiago_team_wr else 0.5
    for player in gw1_expected:
        team = swiss_players[swiss_players["Player"] == player]["Team"].values
        if len(team) > 0:
            wr = santiago_team_wr.get(team[0], 0.5)
            # Very aggressive: top teams get 3x, bottom teams get 0.3x
            # A weak team player scoring 20 in GW1 is worth less than a strong
            # team player scoring 15, because the strong team player will also
            # score in GW2+GW3 without costing a transfer
            longevity = 1.0 + (wr - avg_wr) * 8.0
            longevity = max(0.2, min(3.5, longevity))
            gw1_expected[player] *= longevity

    gw1_team = build_santiago_team(
        swiss_players, gw1_expected, role_map,
        budget=SANTIAGO_BUDGET, n_iter=30000
    )

    santiago_gw_teams = []

    if gw1_team:
        gw1_actual_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week="W1")
        actual_gw1, gw1_scores = score_team_actual(gw1_team, gw1_actual_pts, 2)

        print(f"\n  GW1 Team:")
        print(f"    {'Player':<16} {'Team':<18} {'Slot':>4} {'Role':>4} {'VP':>6} {'ExpB2':>7} {'ActB2':>7}")
        print(f"    {'-'*68}")
        for p in gw1_team["players"]:
            games = gw1_actual_pts.get(p["Player"], [])
            act = compute_best_n_total(games, 2)
            igl = " *IGL*" if p["Player"] == gw1_team["igl"] else ""
            slot = p.get("slot", "?")
            print(f"    {p['Player']:<16} {p['Team']:<18} {slot:>4} {p['role']:>4} "
                  f"{p['price']:>6.1f} {p['expected_best2']:>7.1f} {act:>7.1f}{igl}")
        print(f"    Total VP: {gw1_team['total_vp']:.1f}  |  "
              f"Actual GW1: {actual_gw1:.1f}")
        santiago_gw_teams.append({"label": "Santiago GW1 (Swiss)", "team": gw1_team,
                                  "actual_pts": gw1_actual_pts})
    else:
        print("  FAILED to build GW1 team!")
        santiago_gw_teams.append({"label": "Santiago GW1 (Swiss)", "team": None})

    # --- GW2: Playoffs ---
    print("\n  --- GW2: Playoffs ---")

    # Check which teams are in W2 (from actual data — now known, P?==1 only)
    w2_teams = csv_2026_full[
        (csv_2026_full["Stage"] == "Santiago") & (csv_2026_full["Wk"] == "W2") &
        (csv_2026_full["P?"] == 1)
    ]["Team"].unique().tolist()
    w2_teams = [_normalize_team_name(t) for t in w2_teams]
    print(f"  W2 teams: {', '.join(sorted(set(w2_teams)))}")

    w2_players = santiago_manual[
        santiago_manual["Team"].isin(w2_teams)
    ].copy()
    print(f"  W2 player pool: {len(w2_players)} players")

    # GW1 tournament performances (heavily weighted for GW2 estimation)
    gw1_tournament_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week="W1")

    gw2_expected = compute_santiago_expected_pts(
        w2_players, santiago_hist_games, santiago_team_wr,
        n_games=3, n_best=2,
        tournament_games=gw1_tournament_pts, tournament_weight=0.4
    )

    w2_teams_set = set(w2_teams)
    if gw1_team:
        gw2_result = santiago_transfers(
            gw1_team, w2_players, gw2_expected, role_map,
            budget=SANTIAGO_BUDGET, max_transfers=SANTIAGO_MAX_TRANSFERS, n_iter=10000,
            gw_teams=w2_teams_set
        )
    else:
        gw2_result = build_santiago_team(
            w2_players, gw2_expected, role_map,
            budget=SANTIAGO_BUDGET, n_iter=10000
        )

    if gw2_result:
        gw2_actual_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week="W2")
        actual_gw2, gw2_scores = score_team_actual(gw2_result, gw2_actual_pts, 2)

        print(f"\n  GW2 Team:")
        if gw2_result.get("transfers"):
            for t in gw2_result["transfers"]:
                print(f"    Transfer: {t[0]} -> {t[1]}")
        print(f"    {'Player':<16} {'Team':<18} {'Slot':>4} {'Role':>4} {'VP':>6} {'ExpB2':>7} {'ActB2':>7}")
        print(f"    {'-'*68}")
        for p in gw2_result["players"]:
            games = gw2_actual_pts.get(p["Player"], [])
            act = compute_best_n_total(games, 2)
            igl = " *IGL*" if p["Player"] == gw2_result["igl"] else ""
            slot = p.get("slot", "?")
            print(f"    {p['Player']:<16} {p['Team']:<18} {slot:>4} {p['role']:>4} "
                  f"{p['price']:>6.1f} {p['expected_best2']:>7.1f} {act:>7.1f}{igl}")
        print(f"    Total VP: {gw2_result['total_vp']:.1f}  |  "
              f"Actual GW2: {actual_gw2:.1f}")
        santiago_gw_teams.append({"label": "Santiago GW2 (Playoffs)", "team": gw2_result,
                                  "actual_pts": gw2_actual_pts})
    else:
        print("  FAILED to build GW2 team!")
        santiago_gw_teams.append({"label": "Santiago GW2 (Playoffs)", "team": None})

    # --- GW3: Final Four ---
    print("\n  --- GW3: Final Four ---")

    w3_teams = csv_2026_full[
        (csv_2026_full["Stage"] == "Santiago") & (csv_2026_full["Wk"] == "W3") &
        (csv_2026_full["P?"] == 1)
    ]["Team"].unique().tolist()
    w3_teams = [_normalize_team_name(t) for t in w3_teams]
    print(f"  Final Four teams: {', '.join(sorted(set(w3_teams)))}")

    w3_players = santiago_manual[
        santiago_manual["Team"].isin(w3_teams)
    ].copy()
    print(f"  W3 player pool: {len(w3_players)} players")

    # Check how many games each W3 team plays (for lower bracket risk)
    w3_game_counts = {}
    w3_data = csv_2026_full[
        (csv_2026_full["Stage"] == "Santiago") & (csv_2026_full["Wk"] == "W3") &
        (csv_2026_full["P?"] == 1)
    ]
    w3_teams_set = set(w3_teams)
    for team in w3_teams_set:
        team_games = w3_data[w3_data["Team"] == team]["Game"].nunique()
        w3_game_counts[team] = team_games
    print(f"  Games per team in W3: {w3_game_counts}")

    # GW1+GW2 tournament performances (heavily weighted)
    gw12_tournament_pts = get_player_actual_game_pts(csv_2026_full, "Santiago")
    # Filter to only W1+W2
    gw12_filtered = {}
    for player, games in gw12_tournament_pts.items():
        gw12_filtered[player] = [g for g in games if g["week"] in ["W1", "W2"]]

    gw3_expected = compute_santiago_expected_pts(
        w3_players, santiago_hist_games, santiago_team_wr,
        n_games=2, n_best=2,
        tournament_games=gw12_filtered, tournament_weight=0.5
    )

    # Discount players on teams with only 1 game in W3
    for player in gw3_expected:
        team = w3_players[w3_players["Player"] == player]["Team"].values
        if len(team) > 0:
            team_name = team[0]
            if w3_game_counts.get(team_name, 2) <= 1:
                # Only 1 game means best-2 = just that 1 game
                gw3_expected[player] *= 0.6

    current_gw_team = gw2_result if gw2_result else gw1_team
    w3_teams_set = set(w3_teams)
    if current_gw_team:
        gw3_result = santiago_transfers(
            current_gw_team, w3_players, gw3_expected, role_map,
            budget=SANTIAGO_BUDGET, max_transfers=SANTIAGO_MAX_TRANSFERS, n_iter=30000,
            gw_teams=w3_teams_set
        )
    else:
        gw3_result = build_santiago_team(
            w3_players, gw3_expected, role_map,
            budget=SANTIAGO_BUDGET, n_iter=10000
        )

    if gw3_result:
        gw3_actual_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week="W3")
        actual_gw3, gw3_scores = score_team_actual(gw3_result, gw3_actual_pts, 2)

        print(f"\n  GW3 Team:")
        if gw3_result.get("transfers"):
            for t in gw3_result["transfers"]:
                print(f"    Transfer: {t[0]} -> {t[1]}")
        print(f"    {'Player':<16} {'Team':<18} {'Slot':>4} {'Role':>4} {'VP':>6} {'ExpB2':>7} {'ActB2':>7}")
        print(f"    {'-'*68}")
        for p in gw3_result["players"]:
            games = gw3_actual_pts.get(p["Player"], [])
            act = compute_best_n_total(games, 2)
            igl = " *IGL*" if p["Player"] == gw3_result["igl"] else ""
            slot = p.get("slot", "?")
            print(f"    {p['Player']:<16} {p['Team']:<18} {slot:>4} {p['role']:>4} "
                  f"{p['price']:>6.1f} {p['expected_best2']:>7.1f} {act:>7.1f}{igl}")
        print(f"    Total VP: {gw3_result['total_vp']:.1f}  |  "
              f"Actual GW3: {actual_gw3:.1f}")
        santiago_gw_teams.append({"label": "Santiago GW3 (Final Four)", "team": gw3_result,
                                  "actual_pts": gw3_actual_pts})
    else:
        print("  FAILED to build GW3 team!")
        santiago_gw_teams.append({"label": "Santiago GW3 (Final Four)", "team": None})

    # --- Santiago Total ---
    total_santiago = 0
    for i, td in enumerate(santiago_gw_teams):
        if td["team"]:
            week = f"W{i+1}"
            week_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week=week)
            act, _ = score_team_actual(td["team"], week_pts, 2)
            total_santiago += act

    print(f"\n  TOTAL SANTIAGO POINTS: {total_santiago:.1f}")

    # ================================================================
    #  STEP 5: EXPECTED POINTS TABLES
    # ================================================================
    print("\n" + "=" * 70)
    print("  STEP 5: EXPECTED POINTS (Kickoff + Stage 1)")
    print("=" * 70)

    # --- Kickoff expected pts (best 3 of N games, bootstrap) ---
    print("\n  Computing expected points for all Kickoff players...")
    kickoff_expected_pts = compute_pool_expected_pts(
        kickoff_manual, player_historical_games, kickoff_team_wr,
        n_best=KICKOFF_BEST_N, bracket_type="kickoff"
    )

    # Print top 20 Kickoff players by expected pts with actual comparison
    kickoff_exp_list = []
    for player, exp in kickoff_expected_pts.items():
        actual_games = kickoff_actual_pts.get(player, [])
        actual = compute_best_n_total(actual_games, KICKOFF_BEST_N)
        manual_vp = kickoff_manual[kickoff_manual["Player"] == player]["manual_vp"].values
        manual_vp = manual_vp[0] if len(manual_vp) > 0 else 0
        team = kickoff_manual[kickoff_manual["Player"] == player]["Team"].values
        team = team[0] if len(team) > 0 else "?"
        kickoff_exp_list.append({
            "Player": player, "Team": team, "VP": manual_vp,
            "Expected": exp, "Actual": actual, "Diff": actual - exp
        })
    kickoff_exp_list.sort(key=lambda x: x["Expected"], reverse=True)

    print(f"\n  KICKOFF: Expected vs Actual Best-{KICKOFF_BEST_N} Points (Top 30)")
    print(f"    {'Player':<16} {'Team':<18} {'VP':>5} {'Expected':>9} {'Actual':>7} {'Diff':>7}")
    print(f"    {'-'*65}")
    for p in kickoff_exp_list[:30]:
        diff_str = f"{p['Diff']:+.1f}" if p['Actual'] > 0 else "N/A"
        print(f"    {p['Player']:<16} {p['Team']:<18} {p['VP']:>5.1f} "
              f"{p['Expected']:>9.1f} {p['Actual']:>7.1f} {diff_str:>7}")

    # Correlation of expected vs actual
    exp_vals = [p["Expected"] for p in kickoff_exp_list if p["Actual"] > 0]
    act_vals = [p["Actual"] for p in kickoff_exp_list if p["Actual"] > 0]
    if len(exp_vals) > 5:
        corr = np.corrcoef(exp_vals, act_vals)[0, 1]
        mae = np.mean(np.abs(np.array(exp_vals) - np.array(act_vals)))
        print(f"\n  Kickoff Expected-to-Actual Correlation: {corr:.3f}")
        print(f"  Kickoff Expected-to-Actual MAE: {mae:.1f}")

    # --- Stage 1 expected pts (all 5 games count) ---
    print("\n  Computing expected points for Stage 1 (all 5 gameweeks)...")

    # Training data: everything through Santiago
    stage1_train = load_data_up_to(2026, max_stage="Santiago", include_china=False)
    print(f"  Stage 1 training data: {len(stage1_train)} rows")

    # Build historical games for Stage 1
    stage1_hist_train = stage1_train[stage1_train["P?"] == 1]
    stage1_hist_games = {}
    for (player,), grp in stage1_hist_train.groupby(["Player"]):
        pl = player.lower()
        games = []
        for _, r in grp.iterrows():
            games.append({"pts": r["Pts"], "ppm": r["PPM"]})
        stage1_hist_games[pl] = games

    # Player pool = kickoff prices (the 180 non-China players)
    stage1_players = kickoff_manual.copy()
    # Filter out China teams
    stage1_players = stage1_players[~stage1_players["Team"].isin(CHINA_TEAMS_INTL)]

    # Team win rates from full training data
    stage1_team_wr = compute_team_win_rates(stage1_hist_train)

    # Compute expected total pts: 5 games, all count (n_best=None)
    stage1_expected_pts = compute_pool_expected_pts(
        stage1_players, stage1_hist_games, stage1_team_wr,
        n_best=None, fixed_n_games=5
    )

    # Print top 30
    stage1_exp_list = []
    for player, exp in stage1_expected_pts.items():
        manual_vp = stage1_players[stage1_players["Player"] == player]["manual_vp"].values
        manual_vp = manual_vp[0] if len(manual_vp) > 0 else 0
        team = stage1_players[stage1_players["Player"] == player]["Team"].values
        team = team[0] if len(team) > 0 else "?"
        role = role_map.get(player.lower(), "?")
        stage1_exp_list.append({
            "Player": player, "Team": team, "Role": role,
            "VP": manual_vp, "Expected": exp,
        })
    stage1_exp_list.sort(key=lambda x: x["Expected"], reverse=True)

    print(f"\n  STAGE 1: Expected Total Points (5 Gameweeks) — Top 30")
    print(f"    {'Player':<16} {'Team':<18} {'Role':>4} {'VP':>5} {'Expected':>9}")
    print(f"    {'-'*55}")
    for p in stage1_exp_list[:30]:
        print(f"    {p['Player']:<16} {p['Team']:<18} {p['Role']:>4} "
              f"{p['VP']:>5.1f} {p['Expected']:>9.1f}")

    # Summary stats
    all_exp = [p["Expected"] for p in stage1_exp_list]
    print(f"\n  Stage 1 Expected Points Summary:")
    print(f"    Players: {len(all_exp)}")
    print(f"    Mean: {np.mean(all_exp):.1f}, Median: {np.median(all_exp):.1f}")
    print(f"    Min: {np.min(all_exp):.1f}, Max: {np.max(all_exp):.1f}")

    # ================================================================
    #  STEP 6: CREATE EXCEL FILES
    # ================================================================
    print("\n" + "=" * 70)
    print("  STEP 6: CREATE EXCEL FILES")
    print("=" * 70)

    # Kickoff Excel (single best-3 column + expected pts)
    kickoff_team_data = [{"label": "Kickoff Team (Manual Prices)", "team": kickoff_team,
                          "best_n": KICKOFF_BEST_N}]
    create_intl_excel(
        "Kickoff", kickoff_gen, kickoff_manual, kickoff_actual_pts,
        kickoff_team_data, kickoff_eval,
        expected_pts_dict=kickoff_expected_pts
    )

    # Santiago Excel (3 per-GW columns: Best-2 GW1, Best-2 GW2, Best-2 GW3)
    gw1_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week="W1")
    gw2_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week="W2")
    gw3_pts = get_player_actual_game_pts(csv_2026_full, "Santiago", week="W3")
    santiago_per_gw = [("GW1", gw1_pts), ("GW2", gw2_pts), ("GW3", gw3_pts)]
    create_intl_excel(
        "Santiago", santiago_gen, santiago_manual, santiago_actual_pts,
        santiago_gw_teams, santiago_eval,
        per_gw_pts=santiago_per_gw
    )

    # Stage 1 Expected Points Excel
    stage1_gen = generate_intl_prices(
        stage1_train, stage1_players, all_pickrate_dict, santiago_team_pop,
        stage1_team_wr, target_mean=100 / 11,
        vp_min=6.0, vp_max=15.0, role_map=role_map,
    )
    stage1_eval = {"pickrate_dict": all_pickrate_dict, "best_n": 5}
    # No actual pts yet (Stage 1 hasn't happened), pass empty
    stage1_actual_empty = {}
    create_intl_excel(
        "Stage1_Predictions", stage1_gen, stage1_players, stage1_actual_empty,
        [],  # no team data
        stage1_eval,
        expected_pts_dict=stage1_expected_pts
    )

    print("\n" + "=" * 70)
    print("  ALL DONE!")
    print("=" * 70)


if __name__ == "__main__":
    main()
