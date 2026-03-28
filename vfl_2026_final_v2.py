"""
VFL 2026 Stage 1 — Final Pricing & Team Recommendations (v2).

Generates corrected "Generated + User Input" prices from v3 algo base,
builds optimal team recommendations for all 3 pricing schemes,
creates an Excel workbook, and updates the documentation markdown.

Usage:
    python vfl_2026_final_v2.py
"""
import pandas as pd
import numpy as np
import os
import sys
import warnings
import copy
import textwrap
from datetime import datetime

warnings.filterwarnings("ignore", category=FutureWarning)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
#  Imports from existing codebase
# ---------------------------------------------------------------------------
from pricing_algorithms import (
    load_all_data, compute_player_features, compute_team_win_rates,
    load_pickrate_data, compute_player_pickrates, compute_team_popularity,
    _snap_to_half, CHINA_TEAMS, STAGE_ORDER, GAME_ORDER,
)
from schedule_2026 import (
    SCHEDULE, GROUPS, GW_REGIONS, TEAM_ALIASES,
    get_team_opponent, get_playing_teams, get_team_region,
)

# ---------------------------------------------------------------------------
#  Constants
# ---------------------------------------------------------------------------
VP_MIN, VP_MAX = 5.0, 15.0
BUDGET = 100
SQUAD_SIZE = 11
TARGET_MEAN = 100.0 / 11  # ~9.09
NUM_GWS = 6
MAX_TRANSFERS = 3
RNG = np.random.default_rng(2026)

ROLE_SLOTS = {"D": 2, "C": 2, "I": 2, "S": 2}
WILDCARD_SLOTS = 3

# ---------------------------------------------------------------------------
#  User overrides (from v3 algo feedback)
# ---------------------------------------------------------------------------
USER_OVERRIDES = {
    # Underpriced by algo
    "aspas": 14.5,
    "ara": 12.5,
    "Wo0t": 13.0,
    "Sato": 13.5,
    "Jemkin": 13.0,
    "Miniboo": 12.0,
    "Primmie": 12.5,
    "N4RRATE": 11.0,
    "RieNs": 12.0,
    "Comeback": 12.0,
    "Dantedeu5": 10.0,
    "cNed": 9.5,
    "Pxs": 9.0,
    "Absol": 8.5,
    "Darker": 6.5,
    "something": 14.5,
    "virtyy": 8.5,
    # Overpriced by algo
    "BABYBAY": 10.0,
    "skuba": 10.5,
    "Xross": 8.0,
    "Ivy": 7.5,
    "Francis": 8.5,
    "Dambi": 11.5,
    "Marteen": 11.0,
    "starxo": 7.5,
    "Patmen": 8.0,
    "koalanoob": 7.0,
    "Rb": 8.0,
    "Inspire": 7.0,
    "Free1ng": 8.5,
    "bang": 8.0,
}


# ============================================================================
#  SECTION 1 — DATA LOADING
# ============================================================================

def _read_csv_safe(path):
    try:
        return pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="latin-1")


def _normalize_team_name(name):
    if not isinstance(name, str):
        return name
    for old, new in [
        ("KR\x9a Esports", "KRU Esports"),
        ("KRÜ Esports",    "KRU Esports"),
        ("KR\xdc Esports", "KRU Esports"),
        ("LEVIAT\xb5N",    "LEVIATAN"),
        ("LEVIATÁN",       "LEVIATAN"),
        ("LEVIAT\xc1N",    "LEVIATAN"),
        ("PCIFIC Espor",   "PCIFIC Esports"),
        ("ULF Esports",    "Eternal Fire"),
    ]:
        if name == old:
            return new
    if name in TEAM_ALIASES:
        return TEAM_ALIASES[name]
    return name


def get_team_region_normalized(team):
    """Get region for a team, with normalization."""
    norm = _normalize_team_name(team)
    for region, groups in GROUPS.items():
        for gname, tlist in groups.items():
            for t in tlist:
                if _normalize_team_name(t) == norm or t == team or t == norm:
                    return region
    return None


def load_all_data_with_2026():
    frames = []
    for year, filename in [(2024, "2024 VFL.csv"), (2025, "2025 VFL.csv"),
                           (2026, "2026 VFL.csv")]:
        path = os.path.join(DIR, filename)
        if not os.path.exists(path):
            continue
        df = _read_csv_safe(path)
        if "Team" not in df.columns:
            cols = list(df.columns)
            if len(cols) >= 2:
                cols[0], cols[1] = "Team", "Player"
                df.columns = cols
        df["Year"] = year
        frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    combined["Team"] = combined["Team"].apply(_normalize_team_name)
    combined = combined[~combined["Team"].isin(CHINA_TEAMS)].copy()
    return combined


def load_manual_prices():
    path = os.path.join(DIR, "manual_prices_2026.csv")
    df = _read_csv_safe(path)
    df["Team"] = df["Team"].apply(_normalize_team_name)
    return df


def load_v3_prices():
    path = os.path.join(DIR, "generated_prices_2026_v3.csv")
    df = _read_csv_safe(path)
    df["Team"] = df["Team"].apply(_normalize_team_name)
    return df


def load_pickrate_summary():
    path = os.path.join(DIR, "pickrate_summary.csv")
    if not os.path.exists(path):
        return None
    return pd.read_csv(path)


def build_pickrate_lookup(pickrate_df):
    if pickrate_df is None:
        return {}
    lookup = {}
    for _, row in pickrate_df.iterrows():
        lookup[row["Player"].lower()] = {
            "avg_pickpct": row["avg_pickpct"],
            "avg_rank_pct": row["avg_rank_pct"],
            "events_appeared": row["events_appeared"],
            "total_picks": row["total_picks"],
        }
    return lookup


# ============================================================================
#  SECTION 2 — GENERATE "GENERATED + USER INPUT" PRICES
# ============================================================================

def snap_to_half(val):
    return float(np.clip(round(val * 2) / 2, VP_MIN, VP_MAX))


def generate_corrected_prices(v3_df, manual_df, pickrate_lookup):
    """Generate 'Generated + User Input' prices from v3 algo base.

    Logic:
    1. For players in USER_OVERRIDES: use override price directly
    2. For other players where |v3 - manual| >= 2.0: blend with pickrate
    3. For players where |v3 - manual| < 2.0: use v3 as-is
    4. Snap, clip, iteratively shift to mean ~9.09
    """
    # Build lookups
    gen_lookup = {}
    gen_uncertainty = {}
    for _, row in v3_df.iterrows():
        gen_lookup[row["Player"]] = row["generated_vp"]
        gen_uncertainty[row["Player"]] = row.get("uncertainty", "")

    man_lookup = {}
    region_lookup = {}
    position_lookup = {}
    team_lookup = {}
    for _, row in manual_df.iterrows():
        man_lookup[row["Player"]] = row["Price"]
        region_lookup[row["Player"]] = row.get("Region", "")
        position_lookup[row["Player"]] = row.get("Position", "")
        team_lookup[row["Player"]] = row["Team"]

    corrected = {}
    correction_notes = {}

    for player in manual_df["Player"]:
        algo_price = gen_lookup.get(player)
        manual_price = man_lookup.get(player)

        if algo_price is None:
            corrected[player] = manual_price if manual_price is not None else 9.0
            correction_notes[player] = "manual only (not in algo)"
            continue

        if manual_price is None:
            corrected[player] = algo_price
            correction_notes[player] = "algo only (not in manual)"
            continue

        # Check USER_OVERRIDES first
        if player in USER_OVERRIDES:
            corrected[player] = USER_OVERRIDES[player]
            correction_notes[player] = f"user override: {USER_OVERRIDES[player]}"
            continue

        diff = abs(algo_price - manual_price)

        if diff < 2.0:
            # Agreement zone: use v3 algo price
            corrected[player] = algo_price
            correction_notes[player] = f"v3 algo (diff={diff:.1f} < 2.0)"
        else:
            # Disagreement zone: blend with pickrate adjustment
            pick_info = pickrate_lookup.get(player.lower(), {})
            avg_pickpct = pick_info.get("avg_pickpct", 0)

            # Pickrate adjustment: lean toward algo if popular, manual if unpopular
            if avg_pickpct > 25:
                # Popular: pickrate adjustment leans toward algo
                pickrate_adj = algo_price
            elif avg_pickpct < 10:
                # Unpopular: pickrate adjustment leans toward manual
                pickrate_adj = manual_price
            else:
                # Neutral
                pickrate_adj = (algo_price + manual_price) / 2

            blend = 0.35 * algo_price + 0.45 * manual_price + 0.20 * pickrate_adj
            corrected[player] = blend
            correction_notes[player] = (
                f"blend (pickrate={avg_pickpct:.1f}%): "
                f"v3={algo_price:.1f}, man={manual_price:.1f}"
            )

    # Snap to 0.5, clip
    for player in corrected:
        corrected[player] = snap_to_half(corrected[player])

    # Iteratively adjust mean toward TARGET_MEAN
    players_list = list(corrected.keys())
    for iteration in range(300):
        prices_array = np.array([corrected[p] for p in players_list])
        current_mean = prices_array.mean()
        gap = current_mean - TARGET_MEAN

        if abs(gap) < 0.03:
            break

        if gap > 0:
            # Mean too high: reduce players where corrected > max(algo, manual)
            candidates = []
            for p in players_list:
                algo_p = gen_lookup.get(p, corrected[p])
                man_p = man_lookup.get(p, corrected[p])
                # Prefer reducing players that are above both sources
                excess = corrected[p] - max(algo_p, man_p)
                if corrected[p] > VP_MIN and p not in USER_OVERRIDES:
                    candidates.append((p, excess, corrected[p]))
            candidates.sort(key=lambda x: (-x[1], -x[2]))
            if candidates:
                target_p = candidates[0][0]
                corrected[target_p] = snap_to_half(corrected[target_p] - 0.5)
        else:
            # Mean too low: boost cheapest players where corrected < manual
            candidates = []
            for p in players_list:
                man_p = man_lookup.get(p, corrected[p])
                shortfall = man_p - corrected[p]
                if corrected[p] < VP_MAX and p not in USER_OVERRIDES:
                    candidates.append((p, shortfall, corrected[p]))
            candidates.sort(key=lambda x: (-x[1], x[2]))
            if candidates:
                target_p = candidates[0][0]
                corrected[target_p] = snap_to_half(corrected[target_p] + 0.5)

    # Build result DataFrame
    rows = []
    for player in manual_df["Player"]:
        unc = gen_uncertainty.get(player, "")
        if not isinstance(unc, str):
            unc = str(unc) if pd.notna(unc) else ""
        rows.append({
            "Player": player,
            "Team": team_lookup.get(player, ""),
            "Region": region_lookup.get(player, ""),
            "Position": position_lookup.get(player, ""),
            "corrected_vp": corrected.get(player, 9.0),
            "note": correction_notes.get(player, ""),
            "uncertainty": unc,
        })
    result_df = pd.DataFrame(rows)

    print(f"\n  Generated+User prices for {len(result_df)} players")
    print(f"  Mean: {result_df['corrected_vp'].mean():.3f} (target: {TARGET_MEAN:.3f})")
    print(f"  Median: {result_df['corrected_vp'].median():.2f}")
    print(f"  Range: [{result_df['corrected_vp'].min():.1f}, {result_df['corrected_vp'].max():.1f}]")

    return result_df


# ============================================================================
#  SECTION 3 — TEAM RECOMMENDATIONS (GW1-GW6)
# ============================================================================

def compute_player_avg_ppm(training_df):
    """Compute recency-weighted average PPM per player."""
    result = {}
    for player, grp in training_df.groupby("Player"):
        ppm_vals = grp["PPM"].values
        year_vals = grp["Year"].values
        weights = np.where(year_vals == 2024, 0.5,
                  np.where(year_vals == 2025, 1.0, 2.0))
        if weights.sum() > 0:
            result[player] = np.average(ppm_vals, weights=weights)
        else:
            result[player] = np.mean(ppm_vals)
    return result


def get_player_expected_pts_gw(player, team, gw, player_avg_ppm, team_wr):
    """Estimate expected points for a player in a specific gameweek."""
    playing_teams = get_playing_teams(gw)
    norm_playing = {_normalize_team_name(t) for t in playing_teams}
    norm_team = _normalize_team_name(team)

    if norm_team not in norm_playing:
        return 0.0

    base_ppm = player_avg_ppm.get(player, 3.0)
    avg_maps = 2.3

    # Find opponent
    opp = None
    for t_raw in playing_teams:
        if _normalize_team_name(t_raw) == norm_team:
            opp_raw = get_team_opponent(t_raw, gw)
            if opp_raw:
                opp = _normalize_team_name(opp_raw)
            break

    opp_mult = 1.0
    if opp:
        opp_wr = team_wr.get(opp, 0.5)
        if opp_wr > 0.6:
            opp_mult = 0.85
        elif opp_wr < 0.4:
            opp_mult = 1.15

    return base_ppm * avg_maps * opp_mult


def _check_role_validity(squad):
    """Check if squad meets role constraints: at least 2D, 2C, 2I, 2S."""
    role_counts = {"D": 0, "C": 0, "I": 0, "S": 0}
    for p in squad:
        role = p.get("role", "I")
        if role in role_counts:
            role_counts[role] += 1
    return all(role_counts[r] >= 2 for r in ROLE_SLOTS)


def _build_initial_team(players, gw, budget, n_iter):
    """Build initial team (GW1) using greedy + random restarts."""
    best_team = None
    best_score = -1

    for iteration in range(n_iter):
        noise = RNG.normal(0, 0.3, len(players))
        scored = [(p, p["gw_pts"] / (p["price"] + 0.1) + noise[i])
                  for i, p in enumerate(players)]
        scored.sort(key=lambda x: x[1], reverse=True)

        team = []
        team_vp = 0.0
        team_counts = {}
        role_counts = {"D": 0, "C": 0, "I": 0, "S": 0}
        wildcards_used = 0

        for p, _ in scored:
            if len(team) >= SQUAD_SIZE:
                break

            vp = p["price"]
            t = p["Team"]
            role = p["role"]

            if team_vp + vp > budget:
                continue
            if team_counts.get(t, 0) >= 2:
                continue
            remaining = SQUAD_SIZE - len(team) - 1
            if remaining > 0 and (budget - team_vp - vp) < remaining * VP_MIN:
                continue

            if role_counts.get(role, 0) < ROLE_SLOTS.get(role, 0):
                role_counts[role] += 1
            elif wildcards_used < WILDCARD_SLOTS:
                wildcards_used += 1
            else:
                continue

            team.append(p)
            team_vp += vp
            team_counts[t] = team_counts.get(t, 0) + 1

        if len(team) == SQUAD_SIZE and _check_role_validity(team):
            total_pts = sum(p["gw_pts"] for p in team)
            best_player = max(team, key=lambda p: p["gw_pts"])
            total_with_igl = total_pts + best_player["gw_pts"]

            if total_with_igl > best_score:
                best_score = total_with_igl
                best_team = {
                    "players": list(team),
                    "total_vp": round(team_vp, 2),
                    "expected_pts": round(total_pts, 1),
                    "expected_pts_with_igl": round(total_with_igl, 1),
                    "igl": best_player["Player"],
                    "transfers": [],
                }
    return best_team


def _optimize_transfers(all_players, current_squad, gw, budget,
                        max_transfers, n_iter, amer_bonus=0.0):
    """Find optimal transfers for a gameweek."""
    current_names = {p["Player"] for p in current_squad["players"]}
    current_vp = current_squad["total_vp"]

    player_lookup = {p["Player"]: p for p in all_players}

    # Update current squad with this GW's expected points
    for p in current_squad["players"]:
        if p["Player"] in player_lookup:
            p["gw_pts"] = player_lookup[p["Player"]]["gw_pts"]
        else:
            p["gw_pts"] = 0

    best_result = None
    best_score = -1

    for iteration in range(n_iter):
        squad = [dict(p) for p in current_squad["players"]]
        squad_names = set(current_names)
        squad_vp = current_vp
        transfers_made = []

        n_transfers = int(RNG.integers(0, max_transfers + 1))

        for _ in range(n_transfers):
            squad_pts = [(i, p["gw_pts"]) for i, p in enumerate(squad)]
            squad_pts.sort(key=lambda x: x[1] + float(RNG.normal(0, 2)))
            out_idx = squad_pts[0][0]
            out_player = squad[out_idx]

            candidates = [p for p in all_players
                         if p["Player"] not in squad_names
                         and p["price"] <= budget - squad_vp + out_player["price"]]

            if not candidates:
                continue

            noise = RNG.normal(0, 1.0, len(candidates))
            scored_candidates = []
            for i, c in enumerate(candidates):
                bonus = amer_bonus if c.get("region") == "AMER" else 0
                scored_candidates.append((c, c["gw_pts"] + bonus + float(noise[i])))
            scored_candidates.sort(key=lambda x: x[1], reverse=True)

            replaced = False
            for cand, _ in scored_candidates[:5]:
                new_vp = squad_vp - out_player["price"] + cand["price"]
                if new_vp > budget:
                    continue
                team_count = sum(1 for p in squad if p["Team"] == cand["Team"]
                               and p["Player"] != out_player["Player"])
                if team_count >= 2:
                    continue

                test_squad = [p for p in squad if p["Player"] != out_player["Player"]]
                test_squad.append(cand)

                if not _check_role_validity(test_squad):
                    # Also check with wildcard logic
                    role_counts = {"D": 0, "C": 0, "I": 0, "S": 0}
                    wildcards = 0
                    valid = True
                    for tp in test_squad:
                        role = tp["role"]
                        if role_counts.get(role, 0) < ROLE_SLOTS.get(role, 0):
                            role_counts[role] += 1
                        elif wildcards < WILDCARD_SLOTS:
                            wildcards += 1
                        else:
                            valid = False
                            break
                    if not valid:
                        continue

                squad = test_squad
                squad_names = {p["Player"] for p in squad}
                squad_vp = new_vp
                transfers_made.append((out_player["Player"], cand["Player"]))
                replaced = True
                break

        total_pts = sum(p["gw_pts"] for p in squad)
        if squad:
            best_player = max(squad, key=lambda p: p["gw_pts"])
            total_with_igl = total_pts + best_player["gw_pts"]
        else:
            total_with_igl = 0
            best_player = {"Player": "N/A"}

        if total_with_igl > best_score:
            best_score = total_with_igl
            best_result = {
                "players": [dict(p) for p in squad],
                "total_vp": round(squad_vp, 2),
                "expected_pts": round(total_pts, 1),
                "expected_pts_with_igl": round(total_with_igl, 1),
                "igl": best_player["Player"],
                "transfers": list(transfers_made),
            }

    return best_result


def run_gw_recommendations(prices_df, price_col, player_avg_ppm, team_wr,
                           role_map, label=""):
    """Run GW1-GW6 recommendations for a pricing scheme."""
    print(f"\n  --- {label} GW Recommendations ---")

    all_players = []
    for _, row in prices_df.iterrows():
        region = get_team_region_normalized(row["Team"])
        player_name = row["Player"]
        all_players.append({
            "Player": player_name,
            "Team": row["Team"],
            "price": row[price_col],
            "role": role_map.get(player_name.lower(), row.get("Position", "I")),
            "region": region,
            "gw_pts": 0,
        })

    gw_results = []
    current_squad = None

    for gw in range(1, 7):
        # Compute expected pts for this GW
        for p in all_players:
            p["gw_pts"] = get_player_expected_pts_gw(
                p["Player"], p["Team"], gw, player_avg_ppm, team_wr
            )

        if gw == 1:
            # GW1: EMEA + PAC only
            available = [p for p in all_players
                        if get_team_region_normalized(p["Team"]) in ["EMEA", "PAC"]]
            result = _build_initial_team(available, gw, BUDGET, 5000)
        elif gw == 6:
            # GW6: Need all AMER — overwhelming incentive
            amer_bonus = 20.0
            result = _optimize_transfers(
                all_players, current_squad, gw, BUDGET, MAX_TRANSFERS, 8000,
                amer_bonus=amer_bonus
            )
        else:
            # GW2-5: gradual AMER integration with increasing bonus
            # Must be aggressive enough to have 8+ AMER by GW6
            amer_bonuses = {2: 1.0, 3: 3.0, 4: 5.0, 5: 8.0}
            amer_bonus = amer_bonuses.get(gw, 1.0)
            result = _optimize_transfers(
                all_players, current_squad, gw, BUDGET, MAX_TRANSFERS, 5000,
                amer_bonus=amer_bonus
            )

        if result is None:
            print(f"  GW{gw}: FAILED to find valid team!")
            gw_results.append(None)
            continue

        print(f"\n  GW{gw} ({', '.join(GW_REGIONS[gw])}):")
        if result["transfers"]:
            for out_p, in_p in result["transfers"]:
                print(f"    Transfer: {out_p} -> {in_p}")

        print(f"    {'Player':<18} {'Team':<22} {'VP':>6} {'Role':>4} {'Exp':>7} {'IGL':>4}")
        sorted_p = sorted(result["players"], key=lambda p: p["gw_pts"], reverse=True)
        for p in sorted_p:
            igl = "*" if p["Player"] == result["igl"] else ""
            print(f"    {p['Player']:<18} {p['Team']:<22} {p['price']:>6.1f} "
                  f"{p['role']:>4} {p['gw_pts']:>7.1f} {igl:>4}")
        print(f"    Total VP: {result['total_vp']:.1f}  |  "
              f"Exp: {result['expected_pts']:.1f}  |  "
              f"With IGL: {result['expected_pts_with_igl']:.1f}")

        amer_count = sum(1 for p in result["players"]
                        if get_team_region_normalized(p["Team"]) == "AMER")
        print(f"    AMER players: {amer_count}/11")

        gw_results.append(result)
        current_squad = copy.deepcopy(result)

    return gw_results


# ============================================================================
#  SECTION 4 — CREATE EXCEL FILE
# ============================================================================

def create_excel(manual_df, v3_df, corrected_df,
                 manual_gw, generated_gw, corrected_gw, role_map):
    """Create the Excel workbook with 2 sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Sheet 1: Team Recommendations ----
    ws1 = wb.active
    ws1.title = "Team Recommendations"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill_manual = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_corrected = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_fill_generated = PatternFill(start_color="BF360C", end_color="BF360C", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    gw_font = Font(bold=True, size=11, color="000000")
    gw_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    igl_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    section_cols = ["Player", "Team", "VP", "Role", "Exp Pts", "IGL"]
    section_width = len(section_cols)
    sections = [
        ("MANUAL PRICING", header_fill_manual, manual_gw),
        ("GENERATED + USER", header_fill_corrected, corrected_gw),
        ("GENERATED (ALGO)", header_fill_generated, generated_gw),
    ]

    # Row 1: Section headers
    row = 1
    for sec_idx, (sec_name, sec_fill, _) in enumerate(sections):
        start_col = sec_idx * (section_width + 1) + 1
        cell = ws1.cell(row=row, column=start_col, value=sec_name)
        cell.font = header_font
        cell.fill = sec_fill
        cell.alignment = Alignment(horizontal="center")
        end_col = start_col + section_width - 1
        ws1.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)

    # Row 2: Column sub-headers
    row = 2
    for sec_idx in range(3):
        start_col = sec_idx * (section_width + 1) + 1
        for col_offset, col_name in enumerate(section_cols):
            cell = ws1.cell(row=row, column=start_col + col_offset, value=col_name)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border

    # GW data rows
    row = 3
    for gw_idx in range(6):
        gw_num = gw_idx + 1
        # GW header row
        for sec_idx, (_, _, gw_data) in enumerate(sections):
            start_col = sec_idx * (section_width + 1) + 1
            cell = ws1.cell(row=row, column=start_col,
                           value=f"GW{gw_num} ({', '.join(GW_REGIONS[gw_num])})")
            cell.font = gw_font
            cell.fill = gw_fill
            ws1.merge_cells(start_row=row, start_column=start_col,
                           end_row=row, end_column=start_col + section_width - 1)
        row += 1

        # Transfers row
        for sec_idx, (_, _, gw_data) in enumerate(sections):
            start_col = sec_idx * (section_width + 1) + 1
            result = gw_data[gw_idx] if gw_data and gw_idx < len(gw_data) else None
            if result:
                if result.get("transfers"):
                    out_names = [t[0] for t in result["transfers"]]
                    in_names = [t[1] for t in result["transfers"]]
                    transfers_str = f"OUT {', '.join(out_names)} -> IN {', '.join(in_names)}"
                else:
                    transfers_str = "Initial squad" if gw_idx == 0 else "No transfers"
                cell = ws1.cell(row=row, column=start_col, value=f"Transfers: {transfers_str}")
                cell.font = Font(italic=True, size=9)
                ws1.merge_cells(start_row=row, start_column=start_col,
                               end_row=row, end_column=start_col + section_width - 1)
        row += 1

        # Player rows (11 players)
        for player_idx in range(SQUAD_SIZE):
            for sec_idx, (_, _, gw_data) in enumerate(sections):
                start_col = sec_idx * (section_width + 1) + 1
                result = gw_data[gw_idx] if gw_data and gw_idx < len(gw_data) else None
                if result and player_idx < len(result["players"]):
                    sorted_players = sorted(result["players"],
                                           key=lambda p: p["gw_pts"], reverse=True)
                    p = sorted_players[player_idx]
                    is_igl = p["Player"] == result["igl"]

                    ws1.cell(row=row, column=start_col, value=p["Player"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 1, value=p["Team"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 2, value=p["price"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 3, value=p["role"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 4,
                            value=round(p["gw_pts"], 1)).border = thin_border
                    igl_cell = ws1.cell(row=row, column=start_col + 5,
                                       value="IGL" if is_igl else "")
                    igl_cell.border = thin_border
                    if is_igl:
                        igl_cell.font = Font(bold=True, color="FF0000")
                        igl_cell.fill = igl_fill
            row += 1

        # Totals row
        for sec_idx, (_, sec_fill, gw_data) in enumerate(sections):
            start_col = sec_idx * (section_width + 1) + 1
            result = gw_data[gw_idx] if gw_data and gw_idx < len(gw_data) else None
            if result:
                cell = ws1.cell(row=row, column=start_col, value="TOTAL")
                cell.font = Font(bold=True)
                ws1.cell(row=row, column=start_col + 2,
                        value=result["total_vp"]).font = Font(bold=True)
                ws1.cell(row=row, column=start_col + 4,
                        value=result["expected_pts"]).font = Font(bold=True)
                ws1.cell(row=row, column=start_col + 5,
                        value=f"IGL: {result['expected_pts_with_igl']:.1f}").font = Font(italic=True, size=9)
        row += 2  # blank row

    # Set column widths
    for sec_idx in range(3):
        base = sec_idx * (section_width + 1) + 1
        ws1.column_dimensions[get_column_letter(base)].width = 18
        ws1.column_dimensions[get_column_letter(base + 1)].width = 22
        ws1.column_dimensions[get_column_letter(base + 2)].width = 7
        ws1.column_dimensions[get_column_letter(base + 3)].width = 6
        ws1.column_dimensions[get_column_letter(base + 4)].width = 9
        ws1.column_dimensions[get_column_letter(base + 5)].width = 12
        if sec_idx < 2:
            ws1.column_dimensions[get_column_letter(base + 6)].width = 2

    # ---- Sheet 2: Prices ----
    ws2 = wb.create_sheet("Prices")

    price_headers = ["Player", "Team", "Region", "Position",
                     "Manual VP", "Gen+User VP", "Generated VP", "Uncertainty"]

    for col_idx, header in enumerate(price_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Build price lookups
    gen_price_map = {}
    gen_unc_map = {}
    for _, row_data in v3_df.iterrows():
        gen_price_map[row_data["Player"]] = row_data["generated_vp"]
        unc = row_data.get("uncertainty", "")
        gen_unc_map[row_data["Player"]] = unc if isinstance(unc, str) and pd.notna(unc) else ""

    corr_price_map = {}
    for _, row_data in corrected_df.iterrows():
        corr_price_map[row_data["Player"]] = row_data["corrected_vp"]

    # Sort by Gen+User VP descending
    sorted_players = sorted(manual_df["Player"].tolist(),
                           key=lambda p: corr_price_map.get(p, 0), reverse=True)

    light_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    for data_row_idx, player in enumerate(sorted_players, 2):
        mrow = manual_df[manual_df["Player"] == player].iloc[0]
        manual_vp = mrow["Price"]
        gen_user_vp = corr_price_map.get(player, manual_vp)
        gen_vp = gen_price_map.get(player, manual_vp)
        unc = gen_unc_map.get(player, "")

        ws2.cell(row=data_row_idx, column=1, value=player).border = thin_border
        ws2.cell(row=data_row_idx, column=2, value=mrow["Team"]).border = thin_border
        ws2.cell(row=data_row_idx, column=3,
                value=mrow.get("Region", "")).border = thin_border
        ws2.cell(row=data_row_idx, column=4,
                value=mrow.get("Position", "")).border = thin_border
        c5 = ws2.cell(row=data_row_idx, column=5, value=manual_vp)
        c5.border = thin_border
        c6 = ws2.cell(row=data_row_idx, column=6, value=gen_user_vp)
        c6.border = thin_border
        c7 = ws2.cell(row=data_row_idx, column=7, value=gen_vp)
        c7.border = thin_border
        c8 = ws2.cell(row=data_row_idx, column=8, value=unc)
        c8.border = thin_border

        # Color cells where |Gen+User - Manual| > 2
        if abs(gen_user_vp - manual_vp) > 2:
            c5.fill = light_red
            c6.fill = light_red

    # Column widths for sheet 2
    widths = [18, 22, 8, 8, 10, 12, 12, 30]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    output_path = os.path.join(DIR, "VFL_2026_Stage1_Pricing.xlsx")
    wb.save(output_path)
    print(f"\n  Excel saved to: {output_path}")
    return output_path


# ============================================================================
#  SECTION 5 — UPDATE DOCUMENTATION
# ============================================================================

def create_documentation(manual_df, v3_df, corrected_df):
    """Create updated VFL_2026_PRICING_DOCUMENT.md."""

    gen_map = dict(zip(v3_df["Player"], v3_df["generated_vp"]))
    corr_map = dict(zip(corrected_df["Player"], corrected_df["corrected_vp"]))

    # Build uncertainty lookup
    unc_map = {}
    for _, row in v3_df.iterrows():
        unc = row.get("uncertainty", "")
        if isinstance(unc, str) and "HIGH" in str(unc):
            unc_map[row["Player"]] = "HIGH"
        elif isinstance(unc, str) and "MEDIUM" in str(unc):
            unc_map[row["Player"]] = "MEDIUM"

    # Top 30 comparison table (sorted by Gen+User)
    all_players_sorted = sorted(manual_df["Player"].tolist(),
                                key=lambda p: corr_map.get(p, 0), reverse=True)
    top30 = all_players_sorted[:30]
    table_rows = ""
    for p in top30:
        mrow = manual_df[manual_df["Player"] == p].iloc[0]
        man = mrow["Price"]
        gen_u = corr_map.get(p, man)
        gen = gen_map.get(p, man)
        table_rows += f"| {p} | {mrow['Team']} | {mrow.get('Position','')} | {man:.1f} | {gen_u:.1f} | {gen:.1f} |\n"

    # Build uncertainty table
    high_unc_rows = ""
    med_unc_rows = ""
    for p, level in sorted(unc_map.items(), key=lambda x: x[0]):
        gen_vp = gen_map.get(p, "N/A")
        corr_vp = corr_map.get(p, "N/A")
        team = ""
        match = v3_df[v3_df["Player"] == p]
        if len(match) > 0:
            team = match.iloc[0]["Team"]
        if level == "HIGH":
            high_unc_rows += f"| {p} | {team} | {gen_vp} | {corr_vp} |\n"
        else:
            med_unc_rows += f"| {p} | {team} | {gen_vp} | {corr_vp} |\n"

    # Build corrections table
    override_rows = ""
    for player, price in USER_OVERRIDES.items():
        algo_p = gen_map.get(player, "N/A")
        man_match = manual_df[manual_df["Player"] == player]
        man_p = man_match.iloc[0]["Price"] if len(man_match) > 0 else "N/A"
        direction = "UP" if isinstance(algo_p, (int, float)) and price > algo_p else "DOWN"
        override_rows += f"| {player} | {algo_p} | {man_p} | {price} | {direction} |\n"

    doc = f"""# VFL 2026 Stage 1 Pricing Document

## 1. Executive Summary

This document describes the VFL 2026 Stage 1 pricing system. We built a data-driven
pricing algorithm that generates player prices for 172 professional Valorant players
across the Americas, EMEA, and Pacific regions.

The system produces three price sets:
- **Manual Pricing**: Human-set prices based on expert knowledge of the competitive scene.
- **Generated (Algo v3)**: Pure algorithmic output with adaptive recency weighting,
  outlier dampening, and potential floor.
- **Generated + User Input**: The recommended price set -- v3 algorithm base with human
  corrections for known issues (role changes, limited data, community sentiment).

All prices snap to 0.5 VP increments, range from 5.0 to 15.0 VP, and target a mean
of approximately 9.09 VP (100 budget / 11 players).

---

## 2. How the Algorithm Works

### Overview
The algorithm uses a **Bayesian bootstrap ensemble** approach. In plain terms:

> We ran 100 different versions of the model, each weighting past tournaments
> differently, then averaged the results. This captures uncertainty -- if two events
> disagree about a player's value, the ensemble naturally hedges.

### Training Data
The model trains on all professional VCT match data from 2024 through 2026 Santiago
(Masters 1), covering 14 separate events. More recent events carry higher weight:
2024 Kickoff gets 0.30x weight while 2026 Santiago gets 1.00x.

### Factors and Weights

| Factor | Weight | Description |
|--------|--------|-------------|
| EMA Performance (PPM) | 60% | Core metric: points per map, weighted by recency |
| Team Strength | 12% | Team win rate vs average -- strong teams boost T.Pts |
| Pickrate Popularity | 8% | Historical VFL manager pick rates -- supply/demand |
| Team Brand Popularity | 8% | Fan-favorite teams have inflated demand |
| Opponent Schedule | 7% | Actual 2026 group matchups factored in |
| Consistency Premium | 5% | Low variance in PPM earns a small premium |

### v3 Improvements

#### Adaptive Recency Weighting
Instead of fixed event weights, v3 uses **adaptive recency** that adjusts the decay
rate per player based on their sample size. Players with many games get a steeper
recency curve (recent form matters more), while players with few games get a gentler
curve (we rely more on all available data to avoid overreacting to one bad tournament).

#### Outlier Event Dampening
Single outlier tournaments (either exceptionally good or bad) are dampened to prevent
them from skewing the price. If a player's PPM in one event differs by more than 1.5
standard deviations from their overall mean, that event's weight is reduced by 50%.
This prevents one fluke run from defining a player's price.

#### Potential Floor
For players with limited data (fewer than 10 games), the algorithm applies a
**potential floor** based on their team context and role. This prevents unproven
players on strong teams from being priced at the absolute minimum. The floor is set
at (team average VP - 2.0), capped at 7.0 VP.

#### Pickrate Sentiment Weighting
Historical VFL manager pick rates now directly influence the blending formula. When
the algorithm and manual prices disagree by 2+ VP:
- If a player's pickrate is above 25%, the blend leans toward the algorithm (community
  validates the data-driven price).
- If pickrate is below 10%, the blend leans toward the manual price (community agrees
  the player is mispriced by data alone).

### Role Adjustments
- **Duelists (D)**: PPM discounted by 5% -- they naturally score more kills, so raw PPM
  overstates their fantasy premium vs other roles.
- **Sentinels (S)**: PPM boosted by 5% -- undervalued by raw stats, contribute through
  team play.
- **Controllers (C)**: PPM boosted by 3% -- slight boost for similar reasons.
- **Initiators (I)**: No adjustment (baseline role).

### Distribution Shaping
After computing raw scores, players are ranked by percentile and mapped to a target
S-curve distribution. This ensures a healthy spread of prices rather than clustering
everyone around 8-10 VP.

### Opponent Schedule Awareness
For each team, we look up their 5 group-stage opponents from the actual 2026 Stage 1
schedule. Each opponent's historical win rate determines schedule difficulty:
- Easy schedule (weaker opponents) = slight price boost
- Hard schedule (stronger opponents) = slight price discount

---

## 3. Three Price Sets Explained

### Manual Pricing
Human-set prices based on expert knowledge of the competitive Valorant scene. These
reflect qualitative factors the algorithm cannot see: role changes, team dynamics,
visa issues, player motivation, recent scrimmage performance, etc.

### Generated (Algo v3)
Pure algorithmic output with v3 improvements. Strengths: objective, data-driven,
consistent methodology, handles outliers and recency adaptively. Weaknesses: cannot
account for context it has no data for (roster changes, role swaps, team chemistry).

### Generated + User Input (Recommended)
The best of both worlds. Starts from v3 algorithm base, then applies corrections:

- **Agreement zone** (|v3 - manual| < 2 VP): Uses the v3 algo price as-is.
- **User overrides** ({len(USER_OVERRIDES)} specific players): Expert overrides for
  players where the algo has known blind spots (see Section 4).
- **Disagreement zone** (|v3 - manual| >= 2 VP, no override): Blends prices:
  - 35% v3 algo + 45% manual + 20% pickrate adjustment
  - Pickrate > 25%: adjustment leans toward algo
  - Pickrate < 10%: adjustment leans toward manual

---

## 4. Known Issues and Final Corrections Applied

### User Override Corrections

| Player | Algo v3 | Manual | Final Price | Direction |
|--------|---------|--------|-------------|-----------|
{override_rows}

### General Correction Principles
- For players with limited data: pickrate at old prices is weighted more heavily as
  community sentiment.
- If algo and manual differ by 2+ VP, the corrected price is pulled toward manual but
  never simply copied.
- Must still snap to 0.5 VP increments and target mean ~9.09.

---

## 5. Team Recommendations Strategy

### Schedule Context
- **GW1**: EMEA + PAC only (Americas does not start until GW2)
- **GW2-5**: All three regions play
- **GW6**: Americas ONLY (EMEA and PAC are finished)

### GW-by-GW Strategy

**GW1 (EMEA + PAC)**
Build the strongest possible 11-player squad from EMEA and Pacific players only.

**GW2-5 (All Regions)**
Gradually integrate Americas players using 3 transfers per week maximum.
Americas bonus: GW2=+1, GW3=+2, GW4=+3, GW5=+4 (increasing incentive for early
AMER integration to prepare for GW6).

**GW6 (Americas Only)**
Swap remaining non-Americas players for best available AMER. Only Americas teams play.

### IGL (Captain) Strategy
Each gameweek, the player with the highest expected points is designated as IGL,
doubling their score.

---

## 6. Upsides

- **Data-driven base reduces subjectivity**: The algorithm processes thousands of match
  records objectively.
- **Ensemble approach handles uncertainty**: 100 bootstrap samples produce robust estimates.
- **Schedule-aware pricing**: Actual 2026 group-stage matchups factored in.
- **Pickrate sentiment captures community knowledge**: Historical pick rates reflect
  collective intelligence about player value.
- **Multiple viable team archetypes**: Distribution shaping ensures diverse strategies.
- **v3 outlier dampening**: No single bad/good tournament can wildly skew a price.

---

## 7. Downsides and Risks

- **Limited 2026 data for new/transferred players**: These players have higher uncertainty.
- **Historical performance does not predict roster changes**: Role swaps, coaching changes
  are invisible to the model.
- **0.5 VP granularity creates auto-picks**: Price clustering can force obvious picks.
- **Algorithm needs human override for invisible context**: Meta shifts, scrim results,
  player motivation, visa issues.
- **Pickrate data is retrospective**: Past manager behavior reflects past contexts.
- **Role adjustments are fixed multipliers**: May be too aggressive or conservative for
  individual players.

---

## 8. Price Comparison Table (Top 30 Players by Gen+User Price)

| Player | Team | Pos | Manual VP | Gen+User VP | Generated VP |
|--------|------|-----|-----------|-------------|-------------|
{table_rows}
---

## 9. Uncertainty Flags

Players are flagged with uncertainty levels based on their data availability:

- **HIGH UNCERTAINTY**: Fewer than 3 games in training data, or brand new player with no
  historical data. Prices for these players are estimates and should be reviewed carefully.
- **MEDIUM UNCERTAINTY**: Limited sample (3-6 games) or recent team change. Prices are
  reasonably grounded but have wider confidence intervals.

### HIGH Uncertainty Players

| Player | Team | Algo VP | Gen+User VP |
|--------|------|---------|-------------|
{high_unc_rows if high_unc_rows else "| (none) | | | |\n"}

### MEDIUM Uncertainty Players

| Player | Team | Algo VP | Gen+User VP |
|--------|------|---------|-------------|
{med_unc_rows if med_unc_rows else "| (none) | | | |\n"}

These flags indicate where the algorithm has less confidence. For HIGH uncertainty players,
the Gen+User price may rely more heavily on team context, manual expert input, and community
sentiment (pickrate) than on direct performance data.

---

*Generated on {datetime.now().strftime('%Y-%m-%d')} by VFL Pricing Algorithm v3.0*
"""

    doc_path = os.path.join(DIR, "VFL_2026_PRICING_DOCUMENT.md")
    with open(doc_path, "w", encoding="utf-8") as f:
        f.write(doc)
    print(f"\n  Documentation saved to: {doc_path}")
    return doc_path


# ============================================================================
#  MAIN
# ============================================================================

def main():
    print("=" * 70)
    print("VFL 2026 Stage 1 — Final Pricing & Team Recommendations (v2)")
    print("=" * 70)

    # --- Load data ---
    print("\n[1/5] Loading data...")
    manual_df = load_manual_prices()
    v3_df = load_v3_prices()
    pickrate_df = load_pickrate_summary()
    pickrate_lookup = build_pickrate_lookup(pickrate_df)

    print(f"  Manual prices: {len(manual_df)} players")
    print(f"  V3 algo prices: {len(v3_df)} players")
    print(f"  Pickrate data: {len(pickrate_lookup)} players")

    # Load training data for PPM computation
    all_data = load_all_data_with_2026()
    played_mask = all_data["P?"] == 1
    training_df = all_data[played_mask].copy()
    print(f"  Training data: {len(training_df)} played games, "
          f"{training_df['Player'].nunique()} players")

    team_wr = compute_team_win_rates(training_df)
    player_avg_ppm = compute_player_avg_ppm(training_df)

    # Build role map from manual + v3 data
    role_map = {}
    for _, row in manual_df.iterrows():
        role_map[row["Player"].lower()] = row.get("Position", "I")
    for _, row in v3_df.iterrows():
        role_map[row["Player"].lower()] = row.get("Position", "I")

    # --- Generate corrected prices ---
    print("\n[2/5] Generating 'Generated + User Input' prices...")
    corrected_df = generate_corrected_prices(v3_df, manual_df, pickrate_lookup)

    # --- Build combined prices DataFrame for team recommendations ---
    # We need a single DataFrame with all 3 price columns
    # Start from manual as the base (has all players + metadata)
    combined = manual_df.copy()

    # Add v3 generated prices
    gen_map = dict(zip(v3_df["Player"], v3_df["generated_vp"]))
    combined["generated_vp"] = combined["Player"].map(gen_map).fillna(combined["Price"])

    # Add corrected prices
    corr_map = dict(zip(corrected_df["Player"], corrected_df["corrected_vp"]))
    combined["corrected_vp"] = combined["Player"].map(corr_map).fillna(combined["Price"])

    # --- Run team recommendations for all 3 schemes ---
    print("\n[3/5] Running team recommendations (5000 restarts per GW)...")

    # Manual pricing
    manual_gw = run_gw_recommendations(
        combined, "Price", player_avg_ppm, team_wr, role_map,
        label="Manual"
    )

    # Generated + User pricing
    corrected_gw = run_gw_recommendations(
        combined, "corrected_vp", player_avg_ppm, team_wr, role_map,
        label="Gen+User"
    )

    # Generated (algo v3) pricing
    generated_gw = run_gw_recommendations(
        combined, "generated_vp", player_avg_ppm, team_wr, role_map,
        label="Generated v3"
    )

    # --- Create Excel ---
    print("\n[4/5] Creating Excel workbook...")
    excel_path = create_excel(
        manual_df, v3_df, corrected_df,
        manual_gw, generated_gw, corrected_gw, role_map
    )

    # --- Update documentation ---
    print("\n[5/5] Updating documentation...")
    doc_path = create_documentation(manual_df, v3_df, corrected_df)

    # --- Final verification ---
    print("\n" + "=" * 70)
    print("VERIFICATION")
    print("=" * 70)

    # Check Gen+User mean
    final_mean = corrected_df["corrected_vp"].mean()
    print(f"\n  Gen+User mean: {final_mean:.3f} (target: {TARGET_MEAN:.3f}, "
          f"diff: {abs(final_mean - TARGET_MEAN):.3f})")

    # Check Excel exists
    print(f"  Excel exists: {os.path.exists(excel_path)}")
    print(f"  Doc exists: {os.path.exists(doc_path)}")

    # Check team recommendations
    for label, gw_results in [("Manual", manual_gw), ("Gen+User", corrected_gw),
                               ("Generated", generated_gw)]:
        print(f"\n  {label} recommendations:")
        for gw_idx, result in enumerate(gw_results):
            if result is None:
                print(f"    GW{gw_idx+1}: FAILED")
                continue
            amer_count = sum(1 for p in result["players"]
                           if get_team_region_normalized(p["Team"]) == "AMER")
            # Check no 0-point players in GW6 (wrong region left on squad)
            zero_pts = [p["Player"] for p in result["players"] if p["gw_pts"] == 0]
            status = "OK"
            if zero_pts and gw_idx == 5:  # GW6 only
                status = f"WARNING: {len(zero_pts)} non-AMER players scoring 0"
            elif zero_pts and gw_idx > 0:
                status = f"NOTE: {len(zero_pts)} players scoring 0"
            print(f"    GW{gw_idx+1}: {result['expected_pts_with_igl']:.1f} pts (IGL), "
                  f"AMER={amer_count}/11, VP={result['total_vp']:.1f}, {status}")

    # Distribution summary
    print(f"\n  Gen+User price distribution:")
    buckets = pd.cut(corrected_df["corrected_vp"],
                     bins=[4.5, 6.5, 8.5, 10.5, 12.5, 15.5],
                     labels=["5-6.5", "7-8.5", "9-10.5", "11-12.5", "13-15"])
    print(f"    {buckets.value_counts().sort_index().to_dict()}")

    # Show top overrides applied
    print(f"\n  User overrides applied: {len(USER_OVERRIDES)}")
    for p, target in sorted(USER_OVERRIDES.items(), key=lambda x: -x[1]):
        algo_p = gen_map.get(p, "?")
        final_p = corr_map.get(p, "?")
        print(f"    {p:<18} algo={algo_p}, override={target}, final={final_p}")

    print("\n  DONE.")


if __name__ == "__main__":
    main()
