"""
VFL 2026 Stage 1 — Final Pricing & Team Recommendations.

Generates corrected "Generated + User Input" prices, builds optimal team
recommendations for all 3 pricing schemes (Manual, Generated, Generated+User),
creates an Excel workbook and documentation markdown.

Usage:
    python vfl_2026_final.py
"""
import pandas as pd
import numpy as np
import os
import sys
import warnings
import copy

warnings.filterwarnings("ignore", category=FutureWarning)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
#  Imports from existing codebase
# ---------------------------------------------------------------------------
from pricing_algorithms import (
    load_all_data, compute_player_features, compute_team_win_rates,
    load_pickrate_data, compute_player_pickrates, compute_team_popularity,
    _snap_to_half, CHINA_TEAMS, STAGE_ORDER, GAME_ORDER,
)
from schedule_2026 import (
    SCHEDULE, GROUPS, GW_REGIONS, TEAM_ALIASES,
    get_team_opponent, get_playing_teams, get_team_region,
)

# ---------------------------------------------------------------------------
#  Constants
# ---------------------------------------------------------------------------
VP_MIN, VP_MAX = 5.0, 15.0
BUDGET = 100
SQUAD_SIZE = 11
N_BOOTSTRAP = 100
TARGET_MEAN = 100.0 / 11  # ~9.09
NUM_GWS = 6
MAX_TRANSFERS = 3
RNG = np.random.default_rng(2026)

ROLE_SLOTS = {"D": 2, "C": 2, "I": 2, "S": 2}
WILDCARD_SLOTS = 3

TRAINING_EVENTS = [
    (2024, "Kickoff",   0.30),
    (2024, "Madrid",    0.32),
    (2024, "Stage 1",   0.35),
    (2024, "Shanghai",  0.38),
    (2024, "Stage 2",   0.42),
    (2024, "Champions", 0.48),
    (2025, "Kickoff",   0.55),
    (2025, "bangkok",   0.60),
    (2025, "Stage 1",   0.65),
    (2025, "Toronto",   0.70),
    (2025, "Stage 2",   0.78),
    (2025, "Champions", 0.85),
    (2026, "Kickoff",   0.92),
    (2026, "Santiago",  1.00),
]

TARGET_QUANTILES = np.array([0.0, 0.05, 0.10, 0.20, 0.30, 0.40, 0.50,
                              0.60, 0.70, 0.80, 0.90, 0.95, 1.0])
TARGET_VPS      = np.array([5.0, 5.5,  6.0,  7.0,  7.5,  8.5,  9.0,
                              9.5, 10.0, 11.0, 12.5, 13.5, 15.0])

# ---------------------------------------------------------------------------
#  User corrections for specific players
# ---------------------------------------------------------------------------

# Players the algo OVERPRICES (pull toward manual / lower)
OVERPRICED_CORRECTIONS = {
    # NSH RedForce — peaked for 2 tournaments, not insane all of 2025
    "Xross":   None,  # None = pull toward manual (handled by blend logic)
    "Ivy":     None,
    "Francis": None,
    "Dambi":   None,
    "Rb":      None,
    # BABYBAY & skuba — check pickrate; if low at old price, don't raise
    "BABYBAY": None,
    "skuba":   None,
}

# Players the algo UNDERPRICES — target price range (midpoint used)
UNDERPRICED_CORRECTIONS = {
    "aspas":     (14.5, 15.0),
    "ara":       (12.0, 13.0),
    "Wo0t":      (12.5, 13.0),
    "Sato":      (13.0, 14.0),
    "Jemkin":    (13.0, 13.5),
    "Miniboo":   (12.0, 12.5),
    "Primmie":   (12.5, 13.0),
    "N4RRATE":   (11.0, 11.5),
    "RieNs":     (12.0, 12.5),
    "Comeback":  (11.5, 12.0),
    "Dantedeu5": (9.5, 10.0),
    "cNed":      (9.0, 10.0),
    "Pxs":       (9.0, 9.5),
    "Absol":     (8.0, 9.0),
    "Darker":    (6.5, 7.0),
}


# ============================================================================
#  SECTION 1 — DATA LOADING
# ============================================================================

def _read_csv_safe(path):
    try:
        return pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="latin-1")


def _normalize_team_name(name):
    if not isinstance(name, str):
        return name
    for old, new in [
        ("KR\x9a Esports", "KRU Esports"),
        ("KRÜ Esports",    "KRU Esports"),
        ("KR\xdc Esports", "KRU Esports"),
        ("LEVIAT\xb5N",    "LEVIATAN"),
        ("LEVIATÁN",       "LEVIATAN"),
        ("LEVIAT\xc1N",    "LEVIATAN"),
        ("PCIFIC Espor",   "PCIFIC Esports"),
        ("ULF Esports",    "Eternal Fire"),
    ]:
        if name == old:
            return new
    if name in TEAM_ALIASES:
        return TEAM_ALIASES[name]
    return name


def get_team_region_normalized(team):
    for region, groups in GROUPS.items():
        for gname, tlist in groups.items():
            for t in tlist:
                if _normalize_team_name(t) == team:
                    return region
    return None


def load_all_data_with_2026():
    frames = []
    for year, filename in [(2024, "2024 VFL.csv"), (2025, "2025 VFL.csv"),
                           (2026, "2026 VFL.csv")]:
        path = os.path.join(DIR, filename)
        df = _read_csv_safe(path)
        if "Team" not in df.columns:
            cols = list(df.columns)
            if cols[0].startswith("Unnamed") and cols[1].startswith("Unnamed"):
                cols[0], cols[1] = "Team", "Player"
                df.columns = cols
        df["Year"] = year
        frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    combined["Team"] = combined["Team"].apply(_normalize_team_name)
    combined = combined[~combined["Team"].isin(CHINA_TEAMS)].copy()
    return combined


def load_manual_prices():
    path = os.path.join(DIR, "manual_prices_2026.csv")
    df = _read_csv_safe(path)
    df["Team"] = df["Team"].apply(_normalize_team_name)
    df["Player_lower"] = df["Player"].str.lower()
    return df


def load_generated_prices():
    path = os.path.join(DIR, "generated_prices_2026.csv")
    df = _read_csv_safe(path)
    df["Team"] = df["Team"].apply(_normalize_team_name)
    return df


def load_pickrate_summary():
    path = os.path.join(DIR, "pickrate_summary.csv")
    if not os.path.exists(path):
        return None
    return pd.read_csv(path)


def load_schedule_opponent_strength(team_wr):
    result = {}
    all_teams_in_schedule = set()
    for gw in range(1, 7):
        all_teams_in_schedule |= get_playing_teams(gw)
    for team in all_teams_in_schedule:
        team_n = _normalize_team_name(team)
        opp_wrs = []
        for gw in range(1, 7):
            opp = get_team_opponent(team, gw)
            if opp is not None:
                opp_n = _normalize_team_name(opp)
                wr = team_wr.get(opp_n, 0.5)
                opp_wrs.append(wr)
        result[team_n] = np.mean(opp_wrs) if opp_wrs else 0.5
    return result


# ============================================================================
#  SECTION 2 — GENERATE "GENERATED + USER INPUT" PRICES
# ============================================================================

def snap_to_half(val):
    """Snap to 0.5 VP, clip to [5.0, 15.0]."""
    return float(np.clip(round(val * 2) / 2, VP_MIN, VP_MAX))


def build_pickrate_lookup(pickrate_df):
    """Build case-insensitive pickrate lookup."""
    if pickrate_df is None:
        return {}
    lookup = {}
    for _, row in pickrate_df.iterrows():
        lookup[row["Player"].lower()] = {
            "avg_pickpct": row["avg_pickpct"],
            "avg_rank_pct": row["avg_rank_pct"],
            "events_appeared": row["events_appeared"],
            "total_picks": row["total_picks"],
        }
    return lookup


def generate_corrected_prices(generated_df, manual_df, pickrate_lookup):
    """Generate 'Generated + User Input' prices for each player.

    Logic:
    1. If |algo - manual| < 2.0: use algo price
    2. For specifically named players: apply user corrections
    3. For other 2+ VP disagreements: blend based on pickrate
    4. Snap, clip, then iteratively adjust mean to ~9.09
    """
    # Build lookup from generated prices
    gen_lookup = {}
    for _, row in generated_df.iterrows():
        gen_lookup[row["Player"]] = row["generated_vp"]

    # Build lookup from manual prices
    man_lookup = {}
    for _, row in manual_df.iterrows():
        man_lookup[row["Player"]] = row["Price"]

    # Build region lookup from manual
    region_lookup = {}
    position_lookup = {}
    team_lookup = {}
    for _, row in manual_df.iterrows():
        region_lookup[row["Player"]] = row.get("Region", "")
        position_lookup[row["Player"]] = row.get("Position", "")
        team_lookup[row["Player"]] = row["Team"]

    corrected = {}
    correction_notes = {}

    for player in manual_df["Player"]:
        algo_price = gen_lookup.get(player)
        manual_price = man_lookup.get(player)

        if algo_price is None:
            # Player not in generated prices; use manual
            corrected[player] = manual_price
            correction_notes[player] = "manual only (not in algo)"
            continue

        if manual_price is None:
            corrected[player] = algo_price
            correction_notes[player] = "algo only (not in manual)"
            continue

        diff = abs(algo_price - manual_price)

        # --- Check for specifically named corrections ---
        if player in UNDERPRICED_CORRECTIONS:
            low, high = UNDERPRICED_CORRECTIONS[player]
            target = (low + high) / 2
            corrected[player] = target
            correction_notes[player] = f"user correction: {low}-{high}"
            continue

        if player in OVERPRICED_CORRECTIONS:
            # For overpriced players, pull toward manual price
            # Also check pickrate: if low pickrate at old price, pull more toward manual
            pick_info = pickrate_lookup.get(player.lower(), {})
            avg_pickpct = pick_info.get("avg_pickpct", 0)

            if avg_pickpct < 15 and algo_price > manual_price:
                # Very low pickrate: lean heavily toward manual
                blend = 0.2 * algo_price + 0.8 * manual_price
            elif avg_pickpct < 30 and algo_price > manual_price:
                blend = 0.3 * algo_price + 0.7 * manual_price
            else:
                blend = 0.4 * algo_price + 0.6 * manual_price

            corrected[player] = blend
            correction_notes[player] = (
                f"overpriced correction (pickrate={avg_pickpct:.1f}%): "
                f"algo={algo_price:.1f} -> blend={blend:.1f}"
            )
            continue

        # --- Standard logic ---
        if diff < 2.0:
            # They agree enough, use algo price
            corrected[player] = algo_price
            correction_notes[player] = f"algo (diff={diff:.1f} < 2.0)"
        else:
            # 2+ VP disagreement, blend based on pickrate
            pick_info = pickrate_lookup.get(player.lower(), {})
            avg_pickpct = pick_info.get("avg_pickpct", 0)

            if avg_pickpct > 0:
                # Has pickrate data
                # Determine pickrate-implied direction
                if avg_pickpct > 30 and algo_price < manual_price:
                    # Popular player, algo underprices — boost toward manual
                    pickrate_implied = manual_price
                elif avg_pickpct < 15 and algo_price > manual_price:
                    # Unpopular, algo overprices — reduce toward manual
                    pickrate_implied = manual_price
                else:
                    # Neutral pickrate
                    pickrate_implied = (algo_price + manual_price) / 2

                blend = 0.3 * algo_price + 0.4 * manual_price + 0.3 * pickrate_implied
                correction_notes[player] = (
                    f"blend (pickrate={avg_pickpct:.1f}%): "
                    f"algo={algo_price:.1f}, man={manual_price:.1f}"
                )
            else:
                # No pickrate data
                blend = 0.4 * algo_price + 0.6 * manual_price
                correction_notes[player] = (
                    f"blend (no pickrate): algo={algo_price:.1f}, man={manual_price:.1f}"
                )
            corrected[player] = blend

    # --- Snap to 0.5, clip ---
    for player in corrected:
        corrected[player] = snap_to_half(corrected[player])

    # --- Iteratively adjust mean toward TARGET_MEAN ---
    # Strategy: if mean is too high, find players whose corrected price is above
    # both algo and manual and reduce them by 0.5. If mean too low, boost cheap players.
    players_list = list(corrected.keys())

    for iteration in range(200):
        prices_array = np.array([corrected[p] for p in players_list])
        current_mean = prices_array.mean()
        gap = current_mean - TARGET_MEAN

        if abs(gap) < 0.05:
            break

        if gap > 0:
            # Mean too high: reduce the player where corrected > max(algo, manual)
            # or the one furthest above manual
            candidates = []
            for p in players_list:
                algo_p = gen_lookup.get(p, corrected[p])
                man_p = man_lookup.get(p, corrected[p])
                excess = corrected[p] - max(algo_p, man_p)
                if corrected[p] > VP_MIN:
                    candidates.append((p, excess))
            candidates.sort(key=lambda x: -x[1])
            if candidates:
                target_p = candidates[0][0]
                corrected[target_p] = snap_to_half(corrected[target_p] - 0.5)
        else:
            # Mean too low: boost cheapest player where corrected < manual
            candidates = []
            for p in players_list:
                man_p = man_lookup.get(p, corrected[p])
                shortfall = man_p - corrected[p]
                if corrected[p] < VP_MAX:
                    candidates.append((p, shortfall))
            candidates.sort(key=lambda x: -x[1])
            if candidates:
                target_p = candidates[0][0]
                corrected[target_p] = snap_to_half(corrected[target_p] + 0.5)

    for p in players_list:
        corrected[p] = float(corrected[p])

    # Build result DataFrame
    rows = []
    for player in manual_df["Player"]:
        rows.append({
            "Player": player,
            "Team": team_lookup.get(player, ""),
            "Region": region_lookup.get(player, ""),
            "Position": position_lookup.get(player, ""),
            "corrected_vp": corrected.get(player, 9.0),
            "note": correction_notes.get(player, ""),
        })
    result_df = pd.DataFrame(rows)

    print(f"\n  Generated+User prices for {len(result_df)} players")
    print(f"  Mean: {result_df['corrected_vp'].mean():.2f} (target: {TARGET_MEAN:.2f})")
    print(f"  Median: {result_df['corrected_vp'].median():.2f}")
    print(f"  Range: [{result_df['corrected_vp'].min():.1f}, {result_df['corrected_vp'].max():.1f}]")

    return result_df


# ============================================================================
#  SECTION 3 — TEAM RECOMMENDATIONS (GW1-GW6)
# ============================================================================

def compute_player_avg_ppm(training_df):
    """Compute recency-weighted average PPM per player."""
    result = {}
    for player, grp in training_df.groupby("Player"):
        ppm_vals = grp["PPM"].values
        year_vals = grp["Year"].values
        weights = np.where(year_vals == 2024, 0.5,
                  np.where(year_vals == 2025, 1.0, 2.0))
        if weights.sum() > 0:
            result[player] = np.average(ppm_vals, weights=weights)
        else:
            result[player] = np.mean(ppm_vals)
    return result


def get_player_expected_pts_gw(player, team, gw, player_avg_ppm, team_wr):
    """Estimate expected points for a player in a specific gameweek."""
    playing_teams = get_playing_teams(gw)
    norm_playing = {_normalize_team_name(t) for t in playing_teams}

    if team not in norm_playing:
        return 0.0

    base_ppm = player_avg_ppm.get(player, 3.0)
    avg_maps = 2.3

    # Find opponent
    opp = None
    for t_raw in playing_teams:
        if _normalize_team_name(t_raw) == team:
            opp_raw = get_team_opponent(t_raw, gw)
            if opp_raw:
                opp = _normalize_team_name(opp_raw)
            break

    opp_mult = 1.0
    if opp:
        opp_wr = team_wr.get(opp, 0.5)
        if opp_wr >= 0.6:
            opp_mult = 0.85
        elif opp_wr <= 0.4:
            opp_mult = 1.15
        else:
            opp_mult = 1.0 - (opp_wr - 0.5) * 0.6

    return base_ppm * avg_maps * opp_mult


def _build_initial_team(players, gw, budget, n_iter):
    """Build initial team (GW1) using greedy + random restarts."""
    best_team = None
    best_score = -1

    for iteration in range(n_iter):
        noise = RNG.normal(0, 0.2, len(players))
        scored = [(p, p["gw_pts"] / (p["price"] + 0.1) + noise[i])
                  for i, p in enumerate(players)]
        scored.sort(key=lambda x: x[1], reverse=True)

        team = []
        team_vp = 0.0
        team_counts = {}
        role_counts = {"D": 0, "C": 0, "I": 0, "S": 0}
        wildcards_used = 0

        for p, _ in scored:
            if len(team) >= SQUAD_SIZE:
                break

            vp = p["price"]
            t = p["Team"]
            role = p["role"]

            if team_vp + vp > budget:
                continue
            if team_counts.get(t, 0) >= 2:
                continue
            remaining = SQUAD_SIZE - len(team) - 1
            if remaining > 0 and (budget - team_vp - vp) < remaining * VP_MIN:
                continue

            if role_counts.get(role, 0) < ROLE_SLOTS.get(role, 0):
                role_counts[role] += 1
            elif wildcards_used < WILDCARD_SLOTS:
                wildcards_used += 1
            else:
                continue

            team.append(p)
            team_vp += vp
            team_counts[t] = team_counts.get(t, 0) + 1

        if len(team) == SQUAD_SIZE:
            total_pts = sum(p["gw_pts"] for p in team)
            best_player = max(team, key=lambda p: p["gw_pts"])
            total_with_igl = total_pts + best_player["gw_pts"]

            if total_with_igl > best_score:
                best_score = total_with_igl
                best_team = {
                    "players": list(team),
                    "total_vp": round(team_vp, 2),
                    "expected_pts": round(total_pts, 1),
                    "expected_pts_with_igl": round(total_with_igl, 1),
                    "igl": best_player["Player"],
                    "transfers": [],
                }
    return best_team


def _optimize_transfers(all_players, current_squad, gw, budget,
                        max_transfers, n_iter, amer_bonus=0.0):
    """Find optimal transfers for a gameweek.

    amer_bonus: extra score added for Americas players to incentivize early AMER integration.
    """
    current_names = {p["Player"] for p in current_squad["players"]}
    current_vp = current_squad["total_vp"]

    player_lookup = {p["Player"]: p for p in all_players}

    # Update current squad with this GW's expected points
    for p in current_squad["players"]:
        if p["Player"] in player_lookup:
            p["gw_pts"] = player_lookup[p["Player"]]["gw_pts"]
        else:
            p["gw_pts"] = 0

    best_result = None
    best_score = -1

    for iteration in range(n_iter):
        squad = [dict(p) for p in current_squad["players"]]
        squad_names = set(current_names)
        squad_vp = current_vp
        transfers_made = []

        n_transfers = int(RNG.integers(0, max_transfers + 1))

        for _ in range(n_transfers):
            squad_pts = [(i, p["gw_pts"]) for i, p in enumerate(squad)]
            squad_pts.sort(key=lambda x: x[1] + float(RNG.normal(0, 2)))
            out_idx = squad_pts[0][0]
            out_player = squad[out_idx]

            candidates = [p for p in all_players
                         if p["Player"] not in squad_names
                         and p["price"] <= budget - squad_vp + out_player["price"]]

            if not candidates:
                continue

            noise = RNG.normal(0, 1.0, len(candidates))
            scored_candidates = []
            for i, c in enumerate(candidates):
                bonus = amer_bonus if c.get("region") == "AMER" else 0
                scored_candidates.append((c, c["gw_pts"] + bonus + float(noise[i])))
            scored_candidates.sort(key=lambda x: x[1], reverse=True)

            replaced = False
            for cand, _ in scored_candidates[:5]:
                new_vp = squad_vp - out_player["price"] + cand["price"]
                if new_vp > budget:
                    continue
                team_count = sum(1 for p in squad if p["Team"] == cand["Team"]
                               and p["Player"] != out_player["Player"])
                if team_count >= 2:
                    continue

                test_squad = [p for p in squad if p["Player"] != out_player["Player"]]
                test_squad.append(cand)
                role_counts = {"D": 0, "C": 0, "I": 0, "S": 0}
                wildcards = 0
                valid = True
                for tp in test_squad:
                    role = tp["role"]
                    if role_counts.get(role, 0) < ROLE_SLOTS.get(role, 0):
                        role_counts[role] += 1
                    elif wildcards < WILDCARD_SLOTS:
                        wildcards += 1
                    else:
                        valid = False
                        break
                if not valid:
                    continue

                squad = test_squad
                squad_names = {p["Player"] for p in squad}
                squad_vp = new_vp
                transfers_made.append((out_player["Player"], cand["Player"]))
                replaced = True
                break

        total_pts = sum(p["gw_pts"] for p in squad)
        if squad:
            best_player = max(squad, key=lambda p: p["gw_pts"])
            total_with_igl = total_pts + best_player["gw_pts"]
        else:
            total_with_igl = 0
            best_player = {"Player": "N/A"}

        if total_with_igl > best_score:
            best_score = total_with_igl
            best_result = {
                "players": [dict(p) for p in squad],
                "total_vp": round(squad_vp, 2),
                "expected_pts": round(total_pts, 1),
                "expected_pts_with_igl": round(total_with_igl, 1),
                "igl": best_player["Player"],
                "transfers": list(transfers_made),
            }

    return best_result


def run_gw_recommendations(prices_df, player_avg_ppm, team_wr, role_map, label=""):
    """Run GW1-GW6 recommendations for a pricing scheme.

    Returns list of 6 dicts (one per GW), each with players/transfers/igl/etc.
    """
    print(f"\n  --- {label} GW Recommendations ---")

    # Prepare player list with prices, roles, regions
    all_players = []
    for _, row in prices_df.iterrows():
        region = get_team_region_normalized(row["Team"])
        all_players.append({
            "Player": row["Player"],
            "Team": row["Team"],
            "price": row["price"],
            "role": role_map.get(row["Player"].lower(), "I"),
            "region": region,
            "gw_pts": 0,
        })

    gw_results = []
    current_squad = None

    for gw in range(1, 7):
        playing_teams = get_playing_teams(gw)
        norm_playing = {_normalize_team_name(t) for t in playing_teams}

        # Compute expected pts for this GW
        for p in all_players:
            p["gw_pts"] = get_player_expected_pts_gw(
                p["Player"], p["Team"], gw, player_avg_ppm, team_wr
            )

        if gw == 1:
            # GW1: EMEA + PAC only
            available = [p for p in all_players
                        if get_team_region_normalized(p["Team"]) in ["EMEA", "PAC"]]
            result = _build_initial_team(available, gw, BUDGET, 5000)
        elif gw == 6:
            # GW6: Need all AMER — very strong incentive
            amer_bonus = 10.0  # overwhelming incentive for AMER
            result = _optimize_transfers(
                all_players, current_squad, gw, BUDGET, MAX_TRANSFERS, 8000,
                amer_bonus=amer_bonus
            )
        else:
            # GW2-5: gradual AMER integration with increasing bonus
            # GW2: 1.0, GW3: 2.0, GW4: 3.0, GW5: 5.0
            amer_bonuses = {2: 1.0, 3: 2.0, 4: 3.0, 5: 5.0}
            amer_bonus = amer_bonuses.get(gw, 1.0)
            result = _optimize_transfers(
                all_players, current_squad, gw, BUDGET, MAX_TRANSFERS, 5000,
                amer_bonus=amer_bonus
            )

        if result is None:
            print(f"  GW{gw}: FAILED to find valid team!")
            gw_results.append(None)
            continue

        # Print summary
        print(f"\n  GW{gw} ({', '.join(GW_REGIONS[gw])}):")
        if result["transfers"]:
            for out_p, in_p in result["transfers"]:
                print(f"    Transfer: {out_p} -> {in_p}")

        print(f"    {'Player':<18} {'Team':<22} {'VP':>6} {'Role':>4} {'Exp':>7} {'IGL':>4}")
        sorted_p = sorted(result["players"], key=lambda p: p["gw_pts"], reverse=True)
        for p in sorted_p:
            igl = "*" if p["Player"] == result["igl"] else ""
            print(f"    {p['Player']:<18} {p['Team']:<22} {p['price']:>6.1f} "
                  f"{p['role']:>4} {p['gw_pts']:>7.1f} {igl:>4}")
        print(f"    Total VP: {result['total_vp']:.1f}  |  "
              f"Exp: {result['expected_pts']:.1f}  |  "
              f"With IGL: {result['expected_pts_with_igl']:.1f}")

        amer_count = sum(1 for p in result["players"]
                        if get_team_region_normalized(p["Team"]) == "AMER")
        print(f"    AMER players: {amer_count}/11")

        gw_results.append(result)
        current_squad = copy.deepcopy(result)

    return gw_results


# ============================================================================
#  SECTION 4 — CREATE EXCEL FILE
# ============================================================================

def create_excel(manual_df, generated_df, corrected_df,
                 manual_gw, generated_gw, corrected_gw, role_map):
    """Create the Excel workbook with 2 sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Sheet 1: Team Recommendations ----
    ws1 = wb.active
    ws1.title = "Team Recommendations"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill_manual = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_corrected = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_fill_generated = PatternFill(start_color="BF360C", end_color="BF360C", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    gw_font = Font(bold=True, size=11, color="000000")
    gw_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Column layout: 6 cols per section, 3 sections
    section_cols = ["Player", "Team", "VP", "Role", "Exp Pts", "IGL"]
    section_width = len(section_cols)
    sections = [
        ("MANUAL PRICING", header_fill_manual, manual_gw),
        ("GENERATED + USER", header_fill_corrected, corrected_gw),
        ("GENERATED (ALGO)", header_fill_generated, generated_gw),
    ]

    # Row 1: Section headers
    row = 1
    for sec_idx, (sec_name, sec_fill, _) in enumerate(sections):
        start_col = sec_idx * (section_width + 1) + 1
        cell = ws1.cell(row=row, column=start_col, value=sec_name)
        cell.font = header_font
        cell.fill = sec_fill
        cell.alignment = Alignment(horizontal="center")
        # Merge across section
        end_col = start_col + section_width - 1
        ws1.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)

    # Row 2: Column sub-headers
    row = 2
    for sec_idx in range(3):
        start_col = sec_idx * (section_width + 1) + 1
        for col_offset, col_name in enumerate(section_cols):
            cell = ws1.cell(row=row, column=start_col + col_offset, value=col_name)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border

    # GW data rows
    row = 3
    for gw_idx in range(6):
        gw_num = gw_idx + 1
        # GW header row
        for sec_idx, (_, _, gw_data) in enumerate(sections):
            start_col = sec_idx * (section_width + 1) + 1
            cell = ws1.cell(row=row, column=start_col,
                           value=f"GW{gw_num} ({', '.join(GW_REGIONS[gw_num])})")
            cell.font = gw_font
            cell.fill = gw_fill
            ws1.merge_cells(start_row=row, start_column=start_col,
                           end_row=row, end_column=start_col + section_width - 1)
        row += 1

        # Transfers row
        for sec_idx, (_, _, gw_data) in enumerate(sections):
            start_col = sec_idx * (section_width + 1) + 1
            result = gw_data[gw_idx] if gw_data and gw_idx < len(gw_data) else None
            if result and result.get("transfers"):
                transfers_str = "; ".join(
                    f"{out} -> {inp}" for out, inp in result["transfers"]
                )
                cell = ws1.cell(row=row, column=start_col, value=f"Transfers: {transfers_str}")
                cell.font = Font(italic=True, size=9)
                ws1.merge_cells(start_row=row, start_column=start_col,
                               end_row=row, end_column=start_col + section_width - 1)
        row += 1

        # Player rows (11 players)
        for player_idx in range(SQUAD_SIZE):
            for sec_idx, (_, _, gw_data) in enumerate(sections):
                start_col = sec_idx * (section_width + 1) + 1
                result = gw_data[gw_idx] if gw_data and gw_idx < len(gw_data) else None
                if result and player_idx < len(result["players"]):
                    sorted_players = sorted(result["players"],
                                           key=lambda p: p["gw_pts"], reverse=True)
                    p = sorted_players[player_idx]
                    is_igl = p["Player"] == result["igl"]

                    ws1.cell(row=row, column=start_col, value=p["Player"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 1, value=p["Team"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 2, value=p["price"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 3, value=p["role"]).border = thin_border
                    ws1.cell(row=row, column=start_col + 4,
                            value=round(p["gw_pts"], 1)).border = thin_border
                    igl_cell = ws1.cell(row=row, column=start_col + 5,
                                       value="IGL" if is_igl else "")
                    igl_cell.border = thin_border
                    if is_igl:
                        igl_cell.font = Font(bold=True, color="FF0000")
            row += 1

        # Totals row
        for sec_idx, (_, sec_fill, gw_data) in enumerate(sections):
            start_col = sec_idx * (section_width + 1) + 1
            result = gw_data[gw_idx] if gw_data and gw_idx < len(gw_data) else None
            if result:
                ws1.cell(row=row, column=start_col,
                        value="TOTAL").font = Font(bold=True)
                ws1.cell(row=row, column=start_col + 2,
                        value=result["total_vp"]).font = Font(bold=True)
                ws1.cell(row=row, column=start_col + 4,
                        value=result["expected_pts_with_igl"]).font = Font(bold=True)
                ws1.cell(row=row, column=start_col + 5,
                        value="(w/ IGL)").font = Font(italic=True, size=9)
        row += 2  # blank row

    # Set column widths
    for sec_idx in range(3):
        base = sec_idx * (section_width + 1) + 1
        ws1.column_dimensions[get_column_letter(base)].width = 18      # Player
        ws1.column_dimensions[get_column_letter(base + 1)].width = 22  # Team
        ws1.column_dimensions[get_column_letter(base + 2)].width = 7   # VP
        ws1.column_dimensions[get_column_letter(base + 3)].width = 6   # Role
        ws1.column_dimensions[get_column_letter(base + 4)].width = 9   # Exp Pts
        ws1.column_dimensions[get_column_letter(base + 5)].width = 6   # IGL
        # Gap column
        if sec_idx < 2:
            ws1.column_dimensions[get_column_letter(base + 6)].width = 2

    # ---- Sheet 2: Prices ----
    ws2 = wb.create_sheet("Prices")

    price_headers = ["Player", "Team", "Region", "Position",
                     "Manual VP", "Gen+User VP", "Generated VP",
                     "Diff (Gen+User vs Manual)"]

    # Header row
    for col_idx, header in enumerate(price_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Build merged price data
    gen_price_map = {}
    for _, row_data in generated_df.iterrows():
        gen_price_map[row_data["Player"]] = row_data["generated_vp"]

    corr_price_map = {}
    for _, row_data in corrected_df.iterrows():
        corr_price_map[row_data["Player"]] = row_data["corrected_vp"]

    # Sort by Manual VP descending
    sorted_manual = manual_df.sort_values("Price", ascending=False)

    light_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    light_red = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for data_row_idx, (_, mrow) in enumerate(sorted_manual.iterrows(), 2):
        player = mrow["Player"]
        manual_vp = mrow["Price"]
        gen_user_vp = corr_price_map.get(player, manual_vp)
        gen_vp = gen_price_map.get(player, manual_vp)
        diff = gen_user_vp - manual_vp

        ws2.cell(row=data_row_idx, column=1, value=player).border = thin_border
        ws2.cell(row=data_row_idx, column=2, value=mrow["Team"]).border = thin_border
        ws2.cell(row=data_row_idx, column=3,
                value=mrow.get("Region", "")).border = thin_border
        ws2.cell(row=data_row_idx, column=4,
                value=mrow.get("Position", "")).border = thin_border
        ws2.cell(row=data_row_idx, column=5, value=manual_vp).border = thin_border
        ws2.cell(row=data_row_idx, column=6, value=gen_user_vp).border = thin_border
        ws2.cell(row=data_row_idx, column=7, value=gen_vp).border = thin_border

        diff_cell = ws2.cell(row=data_row_idx, column=8, value=round(diff, 1))
        diff_cell.border = thin_border
        if diff > 0.5:
            diff_cell.fill = light_green
        elif diff < -0.5:
            diff_cell.fill = light_red

    # Column widths for sheet 2
    widths = [18, 22, 8, 8, 10, 12, 12, 22]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Save
    output_path = os.path.join(DIR, "VFL_2026_Stage1_Pricing.xlsx")
    wb.save(output_path)
    print(f"\n  Excel saved to: {output_path}")
    return output_path


# ============================================================================
#  SECTION 5 — CREATE DOCUMENTATION MARKDOWN
# ============================================================================

def create_documentation(manual_df, generated_df, corrected_df):
    """Create VFL_2026_PRICING_DOCUMENT.md."""

    # Build top-30 comparison table
    gen_map = dict(zip(generated_df["Player"], generated_df["generated_vp"]))
    corr_map = dict(zip(corrected_df["Player"], corrected_df["corrected_vp"]))

    top30 = manual_df.sort_values("Price", ascending=False).head(30)
    table_rows = ""
    for _, row in top30.iterrows():
        p = row["Player"]
        man = row["Price"]
        gen_u = corr_map.get(p, man)
        gen = gen_map.get(p, man)
        table_rows += f"| {p} | {row['Team']} | {row.get('Position','')} | {man:.1f} | {gen_u:.1f} | {gen:.1f} |\n"

    doc = f"""# VFL 2026 Stage 1 Pricing Document

## 1. Executive Summary

This document describes the VFL 2026 Stage 1 pricing system. We built a data-driven
pricing algorithm that generates player prices for 172 professional Valorant players
across the Americas, EMEA, and Pacific regions.

The system produces three price sets:
- **Manual Pricing**: Human-set prices based on expert knowledge of the competitive scene.
- **Generated (Algo)**: Pure algorithmic output from a Bayesian bootstrap ensemble model.
- **Generated + User Input**: The recommended price set -- algorithm base with human
  corrections for known issues (role changes, limited data, community sentiment).

All prices snap to 0.5 VP increments, range from 5.0 to 15.0 VP, and target a mean
of approximately 9.09 VP (100 budget / 11 players).

---

## 2. How the Algorithm Works

### Overview
The algorithm uses a **Bayesian bootstrap ensemble** approach. In plain terms:

> We ran 100 different versions of the model, each weighting past tournaments
> differently, then averaged the results. This captures uncertainty -- if two events
> disagree about a player's value, the ensemble naturally hedges.

### Training Data
The model trains on all professional VCT match data from 2024 through 2026 Santiago
(Masters 1), covering 14 separate events. More recent events carry higher weight:
2024 Kickoff gets 0.30x weight while 2026 Santiago gets 1.00x.

### Factors and Weights

| Factor | Weight | Description |
|--------|--------|-------------|
| EMA Performance (PPM) | 60% | Core metric: points per map, weighted by recency |
| Team Strength | 12% | Team win rate vs average -- strong teams boost T.Pts |
| Pickrate Popularity | 8% | Historical VFL manager pick rates -- supply/demand |
| Team Brand Popularity | 8% | Fan-favorite teams have inflated demand |
| Opponent Schedule | 7% | Actual 2026 group matchups factored in |
| Consistency Premium | 5% | Low variance in PPM earns a small premium |

### Role Adjustments
- **Duelists (D)**: PPM discounted by 5% -- they naturally score more kills, so raw PPM
  overstates their fantasy premium vs other roles.
- **Sentinels (S)**: PPM boosted by 5% -- undervalued by raw stats, contribute through
  team play.
- **Controllers (C)**: PPM boosted by 3% -- slight boost for similar reasons.
- **Initiators (I)**: No adjustment (baseline role).

### Distribution Shaping
After computing raw scores, players are ranked by percentile and mapped to a target
S-curve distribution. This ensures a healthy spread of prices rather than clustering
everyone around 8-10 VP.

### Opponent Schedule Awareness
For each team, we look up their 5 group-stage opponents from the actual 2026 Stage 1
schedule. Each opponent's historical win rate determines schedule difficulty:
- Easy schedule (weaker opponents) = slight price boost
- Hard schedule (stronger opponents) = slight price discount

---

## 3. Three Price Sets Explained

### Manual Pricing
Human-set prices based on expert knowledge of the competitive Valorant scene. These
reflect qualitative factors the algorithm cannot see: role changes, team dynamics,
visa issues, player motivation, recent scrimmage performance, etc.

### Generated (Algo)
Pure algorithmic output. Strengths: objective, data-driven, consistent methodology.
Weaknesses: cannot account for context it has no data for (roster changes, role swaps,
team chemistry).

### Generated + User Input (Recommended)
The best of both worlds. Starts from the algorithm's base, then applies human corrections
for known issues:

- **Agreement zone** (|algo - manual| < 2 VP): Uses the algo price -- the two sources
  agree enough that the algo's objectivity is preferred.
- **Disagreement zone** (|algo - manual| >= 2 VP): Blends prices using a weighted formula:
  - 30% algo + 40% manual + 30% pickrate-implied sentiment (if pickrate data exists)
  - 40% algo + 60% manual (if no pickrate data)
- **Specific player corrections**: Expert overrides for players where the algo has known
  blind spots (see Section 4).

---

## 4. Known Issues and Corrections

### Players the Algorithm Overprices

| Player | Team | Issue | Correction |
|--------|------|-------|------------|
| NSH RedForce (Xross, Ivy, Francis, Dambi, Rb) | Nongshim RedForce | Peaked for 2 tournament runs, not consistently elite across all of 2025 | Reduced toward manual prices |
| BABYBAY | G2 Esports | Was 2nd duelist profiting from strong team; now role changes. Low pickrate at previous price (~11 VP) suggests managers agreed he was overvalued | Priced to reflect pickrate sentiment |
| skuba | NRG Esports | Similar to BABYBAY: check pickrate at old prices. If not heavily picked at ~9.5, then 15 is not justified | Reduced based on pickrate sentiment |

### Players the Algorithm Underprices

| Player | Team | Issue | Target Price |
|--------|------|-------|-------------|
| aspas | MIBR | Top 2 player of all time; bad team start does not change that | 14.5-15.0 |
| ara | GIANTX | Definition of inconsistent, just had a bad tourney, capable of 13 VP performance | 12-13 |
| Wo0t | Heretics | Inconsistent but capable of high performance | 12.5-13 |
| Sato | LEVIATAN | Insane last season | 13-14 |
| Jemkin | RRQ | Insane last season | 13-13.5 |
| Miniboo | Team Liquid | Great 2024 and 2025, gets set up well by team | 12-12.5 |
| Primmie | FULL SENSE | Insane performer | 12.5-13 |
| N4RRATE | Sentinels | Was in bad roles, now back as initiator | 11-11.5 |
| RieNs | Heretics | Great player on a great team | 12-12.5 |
| Comeback | Heretics | Star on a good team | 11.5-12 |
| Dantedeu5 | KRU Esports | Great last season; KRU could not play due to visa issues | 9.5-10 |
| cNed | PCIFIC Esports | Legend operator player; duelists should never be that cheap | 9-10 |
| Pxs | LEVIATAN | Duelist for team with good hopes this season | 9-9.5 |
| Absol | ZETA DIVISION | Limited data player | 8-9 |
| Darker | LOUD | Not bad; 5.5 is too low | 6.5-7 |

### General Correction Principles
- For players with limited data: pickrate at old prices is weighted more heavily as
  community sentiment.
- If algo and manual differ by 2+ VP, the corrected price is pulled toward manual but
  never simply copied.
- Must still snap to 0.5 VP increments and target mean ~9.09.

---

## 5. Team Recommendations Strategy

### Schedule Context
- **GW1**: EMEA + PAC only (Americas does not start until GW2)
- **GW2-5**: All three regions play
- **GW6**: Americas ONLY (EMEA and PAC are finished)

### GW-by-GW Strategy

**GW1 (EMEA + PAC)**
Build the strongest possible 11-player squad from EMEA and Pacific players only.
No Americas players are available.

**GW2-5 (All Regions)**
Gradually integrate Americas players using the 3 transfers per week maximum:
- GW2: Bring in 1-3 AMER players (they play AND build toward GW6)
- GW3-4: Continue swapping EMEA/PAC players for AMER players
- GW5: Ensure at least 8 AMER players on roster

The optimizer gives a small bonus to Americas players during GW2-5 to incentivize
early integration, even when an EMEA/PAC player might score slightly more that week.

**GW6 (Americas Only)**
Swap remaining non-Americas players for the best available AMER players. Only Americas
teams play, so non-AMER players score 0 points.

### IGL (Captain) Strategy
Each gameweek, the player with the highest expected points is automatically designated
as IGL, doubling their score. The optimizer accounts for this when evaluating squad
compositions -- a team with one high-ceiling player benefits from the IGL multiplier.

### Transfer Planning
With 3 transfers per week maximum:
- Transfers that bring in Americas players serve double duty: they improve the current
  week AND prepare for GW6.
- The optimizer considers future value by adding an "AMER bonus" that increases each week.

---

## 6. Upsides

- **Data-driven base reduces subjectivity**: The algorithm processes thousands of match
  records objectively rather than relying on memory and gut feeling alone.
- **Ensemble approach handles uncertainty**: 100 bootstrap samples with different event
  weightings produce robust estimates that hedge against any single tournament being an
  outlier.
- **Schedule-aware pricing**: Actual 2026 group-stage matchups are factored into prices,
  not just generic team strength ratings.
- **Pickrate sentiment captures community knowledge**: Historical VFL manager pick rates
  reflect collective intelligence about player value at given prices.
- **Multiple viable team archetypes**: Distribution shaping ensures prices allow top-heavy,
  balanced, and spread team-building strategies rather than forcing one dominant approach.

---

## 7. Downsides and Risks

- **Limited 2026 data for new/transferred players**: Players who changed teams or recently
  joined have less historical data, making their prices more uncertain.
- **Historical performance does not predict roster changes**: The algorithm cannot
  anticipate role swaps, team chemistry issues, or coaching changes.
- **0.5 VP granularity creates auto-picks**: When many players cluster at the same price,
  the one with slightly higher expected points becomes an obvious pick for every manager.
- **Algorithm needs human override for context it cannot see**: Meta shifts, scrim results,
  player motivation, visa issues -- these all affect real performance but are invisible to
  the model.
- **Pickrate data is retrospective**: Past manager behavior reflects past prices and
  contexts, not necessarily future value.
- **Role adjustments are fixed multipliers**: A 5% duelist discount may be too aggressive
  for some duelists and too conservative for others.

---

## 8. Price Comparison Table (Top 30 Players by Manual Price)

| Player | Team | Pos | Manual VP | Gen+User VP | Generated VP |
|--------|------|-----|-----------|-------------|-------------|
{table_rows}
---

*Generated on 2026-03-28 by VFL Pricing Algorithm v2.0*
"""

    output_path = os.path.join(DIR, "VFL_2026_PRICING_DOCUMENT.md")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(doc)
    print(f"\n  Documentation saved to: {output_path}")
    return output_path


# ============================================================================
#  MAIN
# ============================================================================

def main():
    print("=" * 80)
    print("  VFL 2026 FINAL — PRICING & TEAM RECOMMENDATIONS")
    print("=" * 80)

    # ------------------------------------------------------------------
    #  SECTION 1: Load everything
    # ------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("  SECTION 1: DATA LOADING")
    print("=" * 80)

    print("\n  Loading CSV data (2024 + 2025 + 2026)...")
    all_data = load_all_data_with_2026()
    print(f"  Combined: {len(all_data)} rows, {all_data['Player'].nunique()} players")

    print("  Loading manual prices...")
    manual_df = load_manual_prices()
    print(f"  Manual prices: {len(manual_df)} players, mean={manual_df['Price'].mean():.2f}")

    print("  Loading generated prices...")
    generated_df = load_generated_prices()
    print(f"  Generated prices: {len(generated_df)} players, "
          f"mean={generated_df['generated_vp'].mean():.2f}")

    print("  Loading pickrate data...")
    pickrate_df = load_pickrate_summary()
    pickrate_lookup = build_pickrate_lookup(pickrate_df)
    print(f"  Pickrate data: {len(pickrate_lookup)} players")

    # Compute training features for PPM and team win rates
    training_stages = set()
    for year, stage, _ in TRAINING_EVENTS:
        training_stages.add((year, stage))

    train_mask = all_data.apply(
        lambda r: (r["Year"], r["Stage"]) in training_stages and r["P?"] == 1, axis=1
    )
    training_df = all_data[train_mask].copy()
    print(f"  Training data: {len(training_df)} played games")

    team_wr = compute_team_win_rates(training_df)
    player_avg_ppm = compute_player_avg_ppm(training_df)

    # Build role map
    role_map = dict(zip(manual_df["Player_lower"], manual_df["Position"]))

    # ------------------------------------------------------------------
    #  SECTION 2: Generate "Generated + User Input" prices
    # ------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("  SECTION 2: GENERATED + USER INPUT PRICES")
    print("=" * 80)

    corrected_df = generate_corrected_prices(generated_df, manual_df, pickrate_lookup)

    # Print corrections applied
    print("\n  --- Key Corrections Applied ---")
    for _, row in corrected_df.iterrows():
        player = row["Player"]
        gen_vp = generated_df.loc[
            generated_df["Player"] == player, "generated_vp"
        ].values
        gen_vp = gen_vp[0] if len(gen_vp) > 0 else None
        man_vp = manual_df.loc[manual_df["Player"] == player, "Price"].values
        man_vp = man_vp[0] if len(man_vp) > 0 else None
        corr_vp = row["corrected_vp"]

        if gen_vp is not None and man_vp is not None:
            if abs(corr_vp - gen_vp) > 1.0 or abs(corr_vp - man_vp) > 1.0:
                print(f"  {player:<18} Gen={gen_vp:>5.1f}  Man={man_vp:>5.1f}  "
                      f"-> Corrected={corr_vp:>5.1f}  ({row['note']})")

    # ------------------------------------------------------------------
    #  SECTION 3: Team Recommendations for all 3 pricing schemes
    # ------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("  SECTION 3: TEAM RECOMMENDATIONS (GW1-GW6)")
    print("=" * 80)

    # Prepare price DataFrames for optimizer
    # Manual prices
    man_prices = manual_df[["Player", "Team", "Price"]].copy()
    man_prices.rename(columns={"Price": "price"}, inplace=True)

    # Generated prices
    gen_prices = generated_df[["Player", "Team", "generated_vp"]].copy()
    gen_prices.rename(columns={"generated_vp": "price"}, inplace=True)

    # Corrected prices
    corr_prices = corrected_df[["Player", "Team", "corrected_vp"]].copy()
    corr_prices.rename(columns={"corrected_vp": "price"}, inplace=True)

    print("\n  Running Manual Pricing recommendations...")
    manual_gw = run_gw_recommendations(
        man_prices, player_avg_ppm, team_wr, role_map, label="MANUAL PRICING"
    )

    print("\n  Running Generated+User recommendations...")
    corrected_gw = run_gw_recommendations(
        corr_prices, player_avg_ppm, team_wr, role_map, label="GENERATED + USER"
    )

    print("\n  Running Generated (Algo) recommendations...")
    generated_gw = run_gw_recommendations(
        gen_prices, player_avg_ppm, team_wr, role_map, label="GENERATED (ALGO)"
    )

    # ------------------------------------------------------------------
    #  SECTION 4: Create Excel
    # ------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("  SECTION 4: CREATING EXCEL FILE")
    print("=" * 80)

    create_excel(manual_df, generated_df, corrected_df,
                 manual_gw, generated_gw, corrected_gw, role_map)

    # ------------------------------------------------------------------
    #  SECTION 5: Create Documentation
    # ------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("  SECTION 5: CREATING DOCUMENTATION")
    print("=" * 80)

    create_documentation(manual_df, generated_df, corrected_df)

    # ------------------------------------------------------------------
    #  Summary
    # ------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("  DONE!")
    print("=" * 80)
    print(f"  Manual mean:     {manual_df['Price'].mean():.2f}")
    print(f"  Generated mean:  {generated_df['generated_vp'].mean():.2f}")
    print(f"  Corrected mean:  {corrected_df['corrected_vp'].mean():.2f}")
    print(f"  Target mean:     {TARGET_MEAN:.2f}")
    print(f"\n  Files created:")
    print(f"    - VFL_2026_Stage1_Pricing.xlsx")
    print(f"    - VFL_2026_PRICING_DOCUMENT.md")


if __name__ == "__main__":
    main()
