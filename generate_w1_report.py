"""Generate W1 report: GW1 team, VLR results, transfer plans, Excel output."""
import sys
import os
import io
import json

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from pulp import LpMaximize, LpProblem, LpVariable, lpSum, PULP_CBC_CMD, value

from v2.data_loader import load_all_data, load_manual_prices, load_pickrate_summary
from v2.expected_points import calibrate, compute_expected_pts
from v2.pricing import compute_prices
from v2.schedule import GW_REGIONS, get_playing_teams, get_team_opponent, SCHEDULE
from v2.constants import (
    VP_MIN, VP_MAX, BUDGET, SQUAD_SIZE, MAX_TRANSFERS, NUM_GWS,
    ROLE_SLOTS, WILDCARD_SLOTS,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))

# VLR team name -> our team name
VLR_TEAM_ALIASES = {
    "KIWOOM DRX(DRX)": "DRX",
    "DetonatioN FocusMe": "Detonation FocusMe",
}

# User's GW1 team
USER_TEAM = [
    ("D", "xeus"), ("D", "Lar0k"), ("I", "Kushy"), ("I", "Rosé"),
    ("C", "Veqaj"), ("C", "PROFEK"), ("S", "Free1ng"), ("S", "Yetujey"),
    ("W", "Jemkin"), ("W", "Munchkin"), ("W", "BeYN"),
]


def main():
    print("=" * 60)
    print("VFL W1 Report Generator")
    print("=" * 60)

    # --- Load base data ---
    print("\n[1/6] Loading data and calibrating model...")
    all_data = load_all_data()
    cal = calibrate(all_data)
    mp = load_manual_prices()
    pr = load_pickrate_summary()

    train = all_data[(all_data["P?"] == 1)].copy()
    roster = _build_roster(mp)
    ep = compute_expected_pts(train, roster, cal)
    ep = compute_prices(ep, pr)

    manual_map = dict(zip(mp["Player"], mp["Stage1_Price"]))
    ep["ManualVP"] = ep["Player"].map(manual_map)
    prices = dict(zip(ep["Player"], ep["ManualVP"]))

    # --- GW1 optimal team (ManualVP) ---
    print("\n[2/6] Building optimal GW1 team (ManualVP)...")
    gw1_team = optimize_gw1(ep, prices)
    _print_team("Optimal GW1 Team", gw1_team)

    # --- Load W1 actual results ---
    print("\n[3/6] Loading W1 VLR results...")
    w1_pts = load_w1_results(ep)
    _print_w1_summary(w1_pts, ep)

    # --- Update EP with W1 actuals ---
    print("\n[4/6] Updating expected points with W1 actuals...")
    ep_updated = update_ep_with_w1(ep, w1_pts)

    # --- Transfer plans (v2 ILP) ---
    print("\n[5/6] Computing transfer plans (v2 ILP)...")
    opt_transfers = compute_transfer_plan(ep_updated, prices, gw1_team)
    user_gw1 = build_user_gw1(ep_updated, prices)
    user_transfers = compute_transfer_plan(ep_updated, prices, user_gw1)

    # --- v1 transfer plan ---
    print("\n[6/6] Computing v1 transfer plan...")
    v1_ep = compute_v1_ep(train)
    v1_ep_updated = update_ep_with_w1(v1_ep, w1_pts)
    v1_transfers = compute_transfer_plan(v1_ep_updated, prices, user_gw1)

    # --- Generate Excel ---
    print("\nGenerating Excel output...")
    write_excel(ep, ep_updated, w1_pts, gw1_team, opt_transfers,
                user_gw1, user_transfers, v1_transfers, v1_ep_updated, prices)
    print("DONE.")


def _build_roster(mp):
    return {
        r["Player"]: {
            "team": r["Team"], "region": r["Region"], "role": r["Position"],
        }
        for _, r in mp.iterrows()
    }


def optimize_gw1(ep, prices):
    """ILP optimization for GW1 using ManualVP prices."""
    gw1 = ep[ep["GW1"] > 0].copy()
    players = gw1.to_dict("records")
    n = len(players)

    prob = LpProblem("GW1_Manual", LpMaximize)
    x = [LpVariable(f"x_{i}", cat="Binary") for i in range(n)]
    igl = [LpVariable(f"igl_{i}", cat="Binary") for i in range(n)]
    sd = [LpVariable(f"sd_{i}", cat="Binary") for i in range(n)]
    si = [LpVariable(f"si_{i}", cat="Binary") for i in range(n)]
    sc = [LpVariable(f"sc_{i}", cat="Binary") for i in range(n)]
    ss = [LpVariable(f"ss_{i}", cat="Binary") for i in range(n)]
    sw = [LpVariable(f"sw_{i}", cat="Binary") for i in range(n)]

    prob += lpSum(
        x[i] * players[i]["GW1"] + igl[i] * players[i]["GW1"]
        for i in range(n)
    )

    prob += lpSum(x[i] for i in range(n)) == SQUAD_SIZE
    prob += lpSum(
        x[i] * prices.get(players[i]["Player"], 9.0) for i in range(n)
    ) <= BUDGET
    prob += lpSum(igl[i] for i in range(n)) == 1

    for t in set(p["Team"] for p in players):
        prob += lpSum(x[i] for i in range(n) if players[i]["Team"] == t) <= 2

    for i in range(n):
        prob += sd[i] + si[i] + sc[i] + ss[i] + sw[i] == x[i]
        prob += igl[i] <= x[i]
        if players[i]["Role"] != "D": prob += sd[i] == 0
        if players[i]["Role"] != "I": prob += si[i] == 0
        if players[i]["Role"] != "C": prob += sc[i] == 0
        if players[i]["Role"] != "S": prob += ss[i] == 0

    prob += lpSum(sd[i] for i in range(n)) == 2
    prob += lpSum(si[i] for i in range(n)) == 2
    prob += lpSum(sc[i] for i in range(n)) == 2
    prob += lpSum(ss[i] for i in range(n)) == 2
    prob += lpSum(sw[i] for i in range(n)) == 3

    prob.solve(PULP_CBC_CMD(msg=0, timeLimit=60))
    return _extract_gw1(players, x, igl, sd, si, sc, ss, sw, prices, "GW1")


def _extract_gw1(players, x, igl, sd, si, sc, ss, sw, prices, gw_col):
    """Extract GW1 team from solved ILP."""
    team = []
    igl_name = None
    for i, p in enumerate(players):
        if value(x[i]) < 0.5:
            continue
        slot = _get_slot(sd, si, sc, ss, sw, i)
        is_igl = value(igl[i]) > 0.5
        if is_igl:
            igl_name = p["Player"]
        team.append({
            "Player": p["Player"], "Team": p["Team"], "Region": p["Region"],
            "Role": p["Role"], "Slot": slot,
            "VP": prices.get(p["Player"], 9.0),
            "GW1": p[gw_col], "IGL": is_igl,
        })
    team.sort(key=lambda t: "DICSW".index(t["Slot"][0]))
    total_vp = sum(t["VP"] for t in team)
    total_pts = sum(t["GW1"] * (2 if t["IGL"] else 1) for t in team)
    return {"players": team, "total_vp": total_vp, "total_pts": total_pts,
            "igl": igl_name}


def _get_slot(sd, si, sc, ss, sw, i):
    if value(sd[i]) > 0.5: return "D"
    if value(si[i]) > 0.5: return "I"
    if value(sc[i]) > 0.5: return "C"
    if value(ss[i]) > 0.5: return "S"
    return "W"


def load_w1_results(ep):
    """Load scraped W1 results, match to our player database."""
    path = os.path.join(DIR, "data", "w1_vlr_results.json")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    our_players = {p.lower(): p for p in ep["Player"].values}
    our_teams = {t.lower(): t for t in ep["Team"].values}
    w1 = {}

    for key, d in data["player_points"].items():
        name = key.rsplit("_", 1)[0]
        vlr_team = d["team"]
        our_team = VLR_TEAM_ALIASES.get(vlr_team, vlr_team)
        matched = our_players.get(name.lower())
        if matched:
            w1[matched] = {
                "actual_pts": d["pts"], "rating": d["rating"],
                "kills": d["kills"], "deaths": d["deaths"],
                "maps": d["maps"], "team": our_team,
            }
    return w1


def update_ep_with_w1(ep, w1_pts):
    """Blend prior EP with actual W1 performance."""
    df = ep.copy()
    alpha = 0.3  # weight for actual result (1 game = moderate signal)

    for i, row in df.iterrows():
        player = row["Player"]
        if player not in w1_pts:
            continue
        actual = w1_pts[player]["actual_pts"]
        n_maps = w1_pts[player]["maps"]
        prior = row["BasePts"]
        updated = alpha * actual / max(n_maps, 1) * 2.3 + (1 - alpha) * prior
        df.at[i, "BasePts"] = updated

        for gw in range(2, 7):
            col = f"GW{gw}"
            if row[col] > 0:
                ratio = updated / max(prior, 0.1)
                df.at[i, col] = row[col] * ratio

    sv_cols = [f"GW{g}" for g in range(1, 7)]
    df["SeasonValue"] = df[sv_cols].sum(axis=1)
    return df


def build_user_gw1(ep, prices):
    """Build the user's GW1 team structure."""
    team = []
    slot_counts = {"D": 0, "I": 0, "C": 0, "S": 0, "W": 0}
    for slot_type, name in USER_TEAM:
        row = ep[ep["Player"] == name]
        if len(row) == 0:
            continue
        r = row.iloc[0]
        slot_counts[slot_type] += 1
        slot_label = f"{slot_type}{slot_counts[slot_type]}"
        team.append({
            "Player": name, "Team": r["Team"], "Region": r["Region"],
            "Role": r["Role"], "Slot": slot_label,
            "VP": prices.get(name, 9.0), "GW1": r["GW1"], "IGL": False,
        })
    # Set IGL to highest GW1 scorer
    if team:
        best = max(team, key=lambda t: t["GW1"])
        best["IGL"] = True
        igl_name = best["Player"]
    else:
        igl_name = None
    total_vp = sum(t["VP"] for t in team)
    total_pts = sum(t["GW1"] * (2 if t["IGL"] else 1) for t in team)
    return {"players": team, "total_vp": total_vp, "total_pts": total_pts,
            "igl": igl_name}


def compute_transfer_plan(ep, prices, gw1_team):
    """Compute GW2-6 transfer plan using slot-locked ILP."""
    # Build slot roster: list of {slot, player, ...} dicts
    roster = []
    for p in gw1_team["players"]:
        roster.append({
            "slot": p["Slot"], "Player": p["Player"], "Team": p["Team"],
            "Region": p["Region"], "Role": p["Role"],
            "VP": p["VP"],
        })
    plan = [gw1_team]
    for gw in range(2, 7):
        result = _optimize_slot_locked(ep, prices, roster, gw)
        plan.append(result)
        roster = result["roster"]
    return plan


def _slot_role(slot):
    """D1->D, I2->I, W3->None (any role)."""
    ch = slot[0]
    return ch if ch in "DICS" else None


def _optimize_slot_locked(ep, prices, roster, gw):
    """ILP where each slot either keeps its player or swaps 1-for-1.

    Slots are FIXED. Players stay in their slot unless transferred out.
    A transfer replaces one specific slot with a role-eligible player.
    """
    gw_col = f"GW{gw}"
    amer_bonus = {4: 2.0, 5: 5.0}.get(gw, 0.0)

    all_p = ep.to_dict("records")
    p_idx = {r["Player"]: i for i, r in enumerate(all_p)}
    n_all = len(all_p)
    slots = [r["slot"] for r in roster]
    n_slots = len(slots)
    roster_names = {r["Player"] for r in roster}

    def ep_val(j):
        base = all_p[j][gw_col]
        if amer_bonus > 0 and all_p[j].get("Region") == "AMER":
            base += amer_bonus
        return base

    prob = LpProblem(f"GW{gw}_SlotLock", LpMaximize)
    out = [LpVariable(f"out_{s}", cat="Binary") for s in range(n_slots)]
    inp = [[LpVariable(f"in_{s}_{j}", cat="Binary")
            for j in range(n_all)] for s in range(n_slots)]
    igl = [LpVariable(f"igl_{s}", cat="Binary") for s in range(n_slots)]
    slot_ep = [LpVariable(f"sep_{s}", lowBound=0) for s in range(n_slots)]
    igl_ep = [LpVariable(f"iep_{s}", lowBound=0) for s in range(n_slots)]
    BIG_M = 30.0

    # Objective: sum of slot EP + IGL bonus
    prob += lpSum(slot_ep[s] + igl_ep[s] for s in range(n_slots))

    for s in range(n_slots):
        cur_idx = p_idx.get(roster[s]["Player"])
        cur_ep = ep_val(cur_idx) if cur_idx is not None else 0

        # slot_ep = kept_player_ep + transferred_in_player_ep
        prob += slot_ep[s] == (
            (1 - out[s]) * cur_ep
            + lpSum(inp[s][j] * ep_val(j) for j in range(n_all))
        )
        # Exactly one player in if transferred out
        prob += lpSum(inp[s][j] for j in range(n_all)) == out[s]

        # Role eligibility for this slot
        required = _slot_role(slots[s])
        for j in range(n_all):
            if required and all_p[j]["Role"] != required:
                prob += inp[s][j] == 0
            if all_p[j]["Player"] in roster_names:
                prob += inp[s][j] == 0

        # IGL linearization: igl_ep[s] = igl[s] * slot_ep[s]
        prob += igl_ep[s] <= slot_ep[s]
        prob += igl_ep[s] <= BIG_M * igl[s]
        prob += igl_ep[s] >= slot_ep[s] - BIG_M * (1 - igl[s])

    # Exactly 1 IGL
    prob += lpSum(igl[s] for s in range(n_slots)) == 1

    # Max 3 transfers
    prob += lpSum(out[s] for s in range(n_slots)) <= MAX_TRANSFERS

    # Each outside player can only be added to one slot
    for j in range(n_all):
        prob += lpSum(inp[s][j] for s in range(n_slots)) <= 1

    # Budget: total VP of resulting team <= 100
    budget_terms = []
    for s in range(n_slots):
        cur_vp = roster[s]["VP"]
        budget_terms.append((1 - out[s]) * cur_vp)
        for j in range(n_all):
            budget_terms.append(
                inp[s][j] * prices.get(all_p[j]["Player"], 9.0)
            )
    prob += lpSum(budget_terms) <= BUDGET

    # Max 2 per team: for each VCT team, count players on resulting roster
    all_teams = set(r["Team"] for r in roster)
    all_teams |= set(p["Team"] for p in all_p)
    for team in all_teams:
        team_count = []
        for s in range(n_slots):
            if roster[s]["Team"] == team:
                team_count.append(1 - out[s])
            for j in range(n_all):
                if all_p[j]["Team"] == team:
                    team_count.append(inp[s][j])
        if team_count:
            prob += lpSum(team_count) <= 2

    prob.solve(PULP_CBC_CMD(msg=0, timeLimit=60))

    return _extract_slot_result(roster, slots, out, inp, igl, slot_ep,
                                all_p, prices, gw_col, gw)


def _extract_slot_result(roster, slots, out, inp, igl, slot_ep,
                         all_p, prices, gw_col, gw):
    """Extract result from slot-locked ILP."""
    new_roster = []
    team = []
    transfers = []
    igl_name = None

    for s in range(len(slots)):
        is_igl = value(igl[s]) > 0.5
        if value(out[s]) > 0.5:
            # Find who came in
            for j in range(len(all_p)):
                if value(inp[s][j]) > 0.5:
                    p = all_p[j]
                    transfers.append({
                        "out": roster[s]["Player"],
                        "in": p["Player"],
                        "slot": slots[s],
                    })
                    entry = {
                        "slot": slots[s], "Player": p["Player"],
                        "Team": p["Team"], "Region": p["Region"],
                        "Role": p["Role"],
                        "VP": prices.get(p["Player"], 9.0),
                    }
                    new_roster.append(entry)
                    team.append({
                        **entry, "Slot": slots[s],
                        gw_col: p[gw_col], "IGL": is_igl,
                    })
                    if is_igl:
                        igl_name = p["Player"]
                    break
        else:
            r = roster[s]
            new_roster.append(dict(r))
            p_data = next(
                (p for p in all_p if p["Player"] == r["Player"]), None
            )
            ep_gw = p_data[gw_col] if p_data else 0
            team.append({
                "Player": r["Player"], "Team": r["Team"],
                "Region": r["Region"], "Role": r["Role"],
                "Slot": slots[s], "VP": r["VP"],
                gw_col: ep_gw, "IGL": is_igl,
            })
            if is_igl:
                igl_name = r["Player"]

    total_vp = sum(t["VP"] for t in team)
    total_pts = sum(
        t[gw_col] * (2 if t["IGL"] else 1) for t in team
    )
    return {
        "players": team, "total_vp": total_vp, "total_pts": total_pts,
        "igl": igl_name, "transfers": transfers,
        "slots": {t["Player"]: t["Slot"] for t in team},
        "roster": new_roster, "gw": gw,
    }


def compute_v1_ep(train):
    """Compute v1-style expected points (PPM-based with year weights)."""
    mp = load_manual_prices()
    roster = _build_roster(mp)
    played = train[train["P?"] == 1].copy()
    played["PPM"] = pd.to_numeric(played["PPM"], errors="coerce").fillna(0)

    player_avg = {}
    for player, grp in played.groupby("Player"):
        ppm = grp["PPM"].values
        years = grp["Year"].values
        weights = np.where(years == 2024, 0.5,
                  np.where(years == 2025, 1.0, 2.0))
        if weights.sum() > 0:
            player_avg[player] = np.average(ppm, weights=weights)
        else:
            player_avg[player] = np.mean(ppm) if len(ppm) > 0 else 3.0

    rows = []
    for player, info in roster.items():
        ppm = player_avg.get(player, 3.0)
        base = ppm * 2.3
        row = {"Player": player, "Team": info["team"], "Region": info["region"],
               "Role": info["role"], "BasePts": base, "N_Games": 0}
        for gw in range(1, 7):
            playing = get_playing_teams(gw)
            if info["team"] in playing:
                row[f"GW{gw}"] = base
            else:
                row[f"GW{gw}"] = 0.0
        row["SeasonValue"] = sum(row[f"GW{g}"] for g in range(1, 7))
        rows.append(row)

    return pd.DataFrame(rows)


def _print_team(label, team_data):
    print(f"\n  {label}:")
    for p in team_data["players"]:
        gw_pts = p.get("GW1", 0)
        mult = 2 if p["IGL"] else 1
        flag = " ** IGL" if p["IGL"] else ""
        print(f"    {p['Slot']:3s} {p['Player']:15s} {p['Team']:20s} "
              f"VP={p['VP']:5.1f}  EP={gw_pts:5.1f}  Pts={gw_pts*mult:5.1f}{flag}")
    print(f"  Total VP: {team_data['total_vp']:.1f}  |  "
          f"Expected (w/IGL): {team_data['total_pts']:.1f}")


def _print_w1_summary(w1_pts, ep):
    user_names = [n for _, n in USER_TEAM]
    print("\n  User's team W1 actual results:")
    total = 0
    for name in user_names:
        if name in w1_pts:
            d = w1_pts[name]
            print(f"    {name:15s} Pts={d['actual_pts']:5.1f}  "
                  f"R={d['rating']:.2f}  K/D={d['kills']}/{d['deaths']}  "
                  f"Maps={d['maps']}")
            total += d["actual_pts"]
        else:
            print(f"    {name:15s} -- no W1 data")
    print(f"  Total actual W1 pts: {total:.1f}")


def write_excel(ep, ep_updated, w1_pts, gw1_team, opt_plan,
                user_gw1, user_plan, v1_plan, v1_ep, prices):
    """Write everything to a multi-sheet Excel file."""
    wb = Workbook()
    _write_gw1_sheet(wb, gw1_team, "GW1 Optimal Team", w1_pts)
    _write_gw1_sheet(wb, user_gw1, "GW1 User Team", w1_pts)
    _write_w1_results_sheet(wb, w1_pts, ep)
    _write_transfer_sheet(wb, opt_plan, "Transfers (v2 Optimal)")
    _write_transfer_sheet(wb, user_plan, "Transfers (v2 User)")
    _write_transfer_sheet(wb, v1_plan, "Transfers (v1 User)")
    _write_pricing_sheet(wb, ep, prices)
    _write_updated_ep_sheet(wb, ep_updated, w1_pts)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    path = os.path.join(DIR, "output", "VFL_2026_Stage1_W1_Report.xlsx")
    wb.save(path)
    print(f"  Saved: {path}")


def _style_header(ws, row, ncols):
    """Apply header styling to a row."""
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="2F5496", fill_type="solid")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")


def _write_gw1_sheet(wb, team_data, title, w1_pts):
    ws = wb.create_sheet(title)
    headers = ["Slot", "Player", "Team", "Region", "Role", "VP",
               "Exp Pts", "Actual Pts", "Diff", "With IGL", "IGL?"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", fill_type="solid")
    igl_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    actual_total = 0

    for r, p in enumerate(team_data["players"], 2):
        gw_pts = p.get("GW1", 0)
        actual = w1_pts.get(p["Player"], {}).get("actual_pts", "")
        diff = actual - gw_pts if isinstance(actual, (int, float)) else ""
        if isinstance(actual, (int, float)):
            actual_total += actual

        ws.cell(row=r, column=1, value=p["Slot"])
        ws.cell(row=r, column=2, value=p["Player"])
        ws.cell(row=r, column=3, value=p["Team"])
        ws.cell(row=r, column=4, value=p["Region"])
        ws.cell(row=r, column=5, value=p["Role"])
        ws.cell(row=r, column=6, value=p["VP"])
        ws.cell(row=r, column=7, value=round(gw_pts, 1))
        ws.cell(row=r, column=8, value=actual if actual != "" else "N/A")
        ws.cell(row=r, column=9, value=round(diff, 1) if diff != "" else "")
        ws.cell(row=r, column=10, value=round(
            gw_pts * (2 if p["IGL"] else 1), 1))
        ws.cell(row=r, column=11, value="IGL" if p["IGL"] else "")

        if p["IGL"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = igl_fill
        elif isinstance(diff, (int, float)):
            ws.cell(row=r, column=9).fill = green if diff >= 2 else (
                red if diff <= -2 else PatternFill())

    summary_row = len(team_data["players"]) + 3
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=team_data["total_vp"])
    ws.cell(row=summary_row, column=8, value=actual_total)
    ws.cell(row=summary_row, column=10, value=round(team_data["total_pts"], 1))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _write_w1_results_sheet(wb, w1_pts, ep):
    ws = wb.create_sheet("W1 Actual Results")
    headers = ["Player", "Team", "Region", "Role", "VP",
               "Actual Pts", "Expected Pts", "Diff",
               "Rating", "Kills", "Deaths", "Maps"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    rows_data = []
    for _, r in ep.iterrows():
        name = r["Player"]
        if name not in w1_pts:
            continue
        w = w1_pts[name]
        rows_data.append({
            "Player": name, "Team": r["Team"], "Region": r["Region"],
            "Role": r["Role"], "VP": r["ManualVP"],
            "Actual": w["actual_pts"], "Expected": r["GW1"],
            "Diff": w["actual_pts"] - r["GW1"],
            "Rating": w["rating"], "Kills": w["kills"],
            "Deaths": w["deaths"], "Maps": w["maps"],
        })

    rows_data.sort(key=lambda x: x["Actual"], reverse=True)
    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", fill_type="solid")

    for row_idx, d in enumerate(rows_data, 2):
        ws.cell(row=row_idx, column=1, value=d["Player"])
        ws.cell(row=row_idx, column=2, value=d["Team"])
        ws.cell(row=row_idx, column=3, value=d["Region"])
        ws.cell(row=row_idx, column=4, value=d["Role"])
        ws.cell(row=row_idx, column=5, value=d["VP"])
        ws.cell(row=row_idx, column=6, value=d["Actual"])
        ws.cell(row=row_idx, column=7, value=round(d["Expected"], 1))
        ws.cell(row=row_idx, column=8, value=round(d["Diff"], 1))
        ws.cell(row=row_idx, column=9, value=d["Rating"])
        ws.cell(row=row_idx, column=10, value=d["Kills"])
        ws.cell(row=row_idx, column=11, value=d["Deaths"])
        ws.cell(row=row_idx, column=12, value=d["Maps"])

        diff_cell = ws.cell(row=row_idx, column=8)
        diff_cell.fill = green if d["Diff"] >= 2 else red if d["Diff"] <= -2 else PatternFill()

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _write_transfer_sheet(wb, plan, title):
    ws = wb.create_sheet(title)
    row = 1

    for gw_idx, gw_data in enumerate(plan):
        gw = gw_idx + 1
        ws.cell(row=row, column=1, value=f"GW{gw}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        if gw > 1 and gw_data.get("transfers"):
            for t in gw_data["transfers"]:
                ws.cell(row=row, column=1, value="TRANSFER:")
                ws.cell(row=row, column=2, value=f"OUT: {t.get('out', '?')}")
                ws.cell(row=row, column=3, value=f"IN: {t['in']}")
                ws.cell(row=row, column=4, value=f"Slot: {t['slot']}")
                for c in range(1, 5):
                    ws.cell(row=row, column=c).fill = PatternFill(
                        start_color="DDEBF7", fill_type="solid"
                    )
                row += 1

        headers = ["Slot", "Player", "Team", "Region", "VP",
                   "Exp Pts", "With IGL", "IGL?"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1

        gw_col = f"GW{gw}"
        for p in gw_data["players"]:
            pts = p.get(gw_col, p.get("GW1", 0))
            ws.cell(row=row, column=1, value=p.get("Slot", "?"))
            ws.cell(row=row, column=2, value=p["Player"])
            ws.cell(row=row, column=3, value=p["Team"])
            ws.cell(row=row, column=4, value=p["Region"])
            ws.cell(row=row, column=5, value=p["VP"])
            ws.cell(row=row, column=6, value=round(pts, 1))
            ws.cell(row=row, column=7, value=round(pts * (2 if p.get("IGL") else 1), 1))
            ws.cell(row=row, column=8, value="IGL" if p.get("IGL") else "")
            row += 1

        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=5, value=round(gw_data["total_vp"], 1))
        ws.cell(row=row, column=7, value=round(gw_data["total_pts"], 1))
        row += 2

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _write_pricing_sheet(wb, ep, prices):
    ws = wb.create_sheet("Pricing")
    headers = ["#", "Player", "Team", "Region", "Role",
               "Manual VP", "Suggested VP", "Diff",
               "GW1", "GW2", "GW3", "GW4", "GW5", "GW6", "Season Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    sorted_ep = ep.sort_values("SuggestedVP", ascending=False)
    for rank, (_, r) in enumerate(sorted_ep.iterrows(), 1):
        row = rank + 1
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=r["Player"])
        ws.cell(row=row, column=3, value=r["Team"])
        ws.cell(row=row, column=4, value=r["Region"])
        ws.cell(row=row, column=5, value=r["Role"])
        ws.cell(row=row, column=6, value=r["ManualVP"])
        ws.cell(row=row, column=7, value=r["SuggestedVP"])
        ws.cell(row=row, column=8, value=round(r["SuggestedVP"] - r["ManualVP"], 1))
        for gw in range(1, 7):
            ws.cell(row=row, column=8 + gw, value=round(r[f"GW{gw}"], 1))
        ws.cell(row=row, column=15, value=round(r["SeasonValue"], 1))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13


def _write_updated_ep_sheet(wb, ep_updated, w1_pts):
    ws = wb.create_sheet("Updated EP (Post-W1)")
    headers = ["Player", "Team", "Region", "Role",
               "Base Pts (Updated)", "GW2", "GW3", "GW4", "GW5", "GW6",
               "Remaining Value", "W1 Actual"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    sorted_ep = ep_updated.sort_values("BasePts", ascending=False)
    for rank, (_, r) in enumerate(sorted_ep.iterrows(), 1):
        row = rank + 1
        ws.cell(row=row, column=1, value=r["Player"])
        ws.cell(row=row, column=2, value=r["Team"])
        ws.cell(row=row, column=3, value=r["Region"])
        ws.cell(row=row, column=4, value=r["Role"])
        ws.cell(row=row, column=5, value=round(r["BasePts"], 2))
        for gw in range(2, 7):
            ws.cell(row=row, column=4 + gw - 1, value=round(r[f"GW{gw}"], 1))
        remaining = sum(r[f"GW{g}"] for g in range(2, 7))
        ws.cell(row=row, column=11, value=round(remaining, 1))
        w1_actual = w1_pts.get(r["Player"], {}).get("actual_pts", "")
        ws.cell(row=row, column=12, value=w1_actual if w1_actual != "" else "N/A")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


if __name__ == "__main__":
    main()
